"""Affiliate managing endpoints."""
import json
import logging
import uuid
from datetime import datetime, timedelta
from typing import Annotated, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from fastapi.responses import RedirectResponse
from pydantic import BaseModel
from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.v1.endpoints.auth import get_current_user, get_shared_cafe24_user
from app.core.config import get_settings
from app.db.database import get_db
from app.models.affiliate import (
    AffiliateCampaign,
    AffiliatePartner,
    AffiliateSettlement,
    ReferralClick,
    ReferralConversion,
    ReferralProgram,
)
from app.models.partner_campaign import PartnerCampaign
from app.models.points import PointTransaction
from app.models.user import User

_settings = get_settings()

logger = logging.getLogger(__name__)

router = APIRouter()


async def _resolve_cafe24_user(current_user: User, db: AsyncSession) -> User:
    """현재 유저가 Cafe24 토큰 없으면 공유 유저 반환. 둘 다 없으면 400."""
    if current_user.cafe24_access_token:
        return current_user
    shared = await get_shared_cafe24_user(db)
    if shared:
        return shared
    raise HTTPException(status_code=400, detail="Cafe24 스토어 연결이 필요합니다.")


# ---------------------------------------------------------------------------
# Pydantic schemas
# ---------------------------------------------------------------------------

class CampaignCreate(BaseModel):
    name: str
    product: str = ""
    description: Optional[str] = None
    commission_type: str = "percentage"
    commission_rate: float = 10.0
    status: str = "active"
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    landing_url: Optional[str] = None
    # Cafe24 단일 상품 연결 (Phase 2 — 백워드 호환)
    cafe24_product_no: Optional[int] = None
    discount_type: Optional[str] = None   # percentage | fixed | shipping
    discount_value: Optional[float] = None
    # Phase 6 — 다중 상품 + 비공개 카테고리 자동 생성
    cafe24_product_nos: Optional[List[int]] = None
    auto_create_category: bool = False
    cafe24_category_name: Optional[str] = None
    cafe24_category_parent_no: Optional[int] = 1


class CampaignUpdate(BaseModel):
    name: Optional[str] = None
    product: Optional[str] = None
    description: Optional[str] = None
    commission_type: Optional[str] = None
    commission_rate: Optional[float] = None
    status: Optional[str] = None
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    landing_url: Optional[str] = None
    discount_type: Optional[str] = None
    discount_value: Optional[float] = None
    cafe24_product_nos: Optional[List[int]] = None  # 카테고리 상품 재구성용


class PartnerCreate(BaseModel):
    campaign_id: Optional[int] = None
    campaign_ids: List[int] = []  # Phase 5 M:N
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    channel: str = "instagram"
    channels: Optional[List[str]] = None  # multi-channel support
    followers: int = 0
    memo: Optional[str] = None
    # 활동 그룹 — crew(크루) / gongu(공구) / ad(광고) / other(기타). 기본 crew
    partner_group: Optional[str] = None


class PartnerUpdate(BaseModel):
    campaign_id: Optional[int] = None
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    channel: Optional[str] = None
    channels: Optional[List[str]] = None  # multi-channel support
    followers: Optional[int] = None
    memo: Optional[str] = None
    referral_link: Optional[str] = None
    partner_group: Optional[str] = None


class SettlementCreate(BaseModel):
    partner_id: int
    amount: float
    period_start: Optional[datetime] = None
    period_end: Optional[datetime] = None


class ReferralProgramCreate(BaseModel):
    name: str
    reward_type: str = "points"
    referrer_reward: float = 0
    referee_reward: float = 0
    max_rewards_per_user: Optional[int] = None
    status: str = "active"


class ReferralProgramUpdate(BaseModel):
    name: Optional[str] = None
    reward_type: Optional[str] = None
    referrer_reward: Optional[float] = None
    referee_reward: Optional[float] = None
    max_rewards_per_user: Optional[int] = None
    status: Optional[str] = None


class ConversionCreate(BaseModel):
    referral_code: str
    order_id: Optional[str] = None
    order_amount: float = 0
    cookie_id: Optional[str] = None


class PartnerCampaignAdd(BaseModel):
    campaign_id: int


class PointAwardRequest(BaseModel):
    user_id: int
    amount: float
    reason: str = "manual"
    memo: Optional[str] = None


# ---------------------------------------------------------------------------
# Campaign CRUD
# ---------------------------------------------------------------------------

@router.post("/campaigns", status_code=status.HTTP_201_CREATED)
async def create_campaign(
    payload: CampaignCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    어필리에이트 캠페인 생성.

    분기:
      A) cafe24_product_no 단독: Phase 2 단일상품 캠페인 (백워드 호환)
      B) cafe24_product_nos[] + auto_create_category: Phase 6 비공개 카테고리 캠페인
         - 카테고리 자동 생성 → 상품들 attach → 카테고리 단위 쿠폰 발급
    """
    import app.services.cafe24 as cafe24_svc

    data = payload.model_dump()
    cafe24_product_no = data.pop("cafe24_product_no", None)
    cafe24_product_nos = data.pop("cafe24_product_nos", None) or []
    auto_create_category = data.pop("auto_create_category", False)
    cafe24_category_name = data.pop("cafe24_category_name", None)
    cafe24_category_parent_no = data.pop("cafe24_category_parent_no", 1) or 1
    discount_type = data.pop("discount_type", None)
    discount_value = data.pop("discount_value", None)

    # timezone-aware → naive 변환
    for key in ('start_date', 'end_date'):
        if data.get(key) and hasattr(data[key], 'replace'):
            data[key] = data[key].replace(tzinfo=None)

    coupon_warning: Optional[str] = None
    category_warning: Optional[str] = None

    # 다중 상품 모드 판정: product_nos가 1개여도 auto_create_category=True면 카테고리 모드
    use_category_mode = bool(auto_create_category and cafe24_product_nos)

    if use_category_mode:
        # ─── B) 비공개 카테고리 캠페인 ────────────────────────────────────────
        cafe24_user = await _resolve_cafe24_user(current_user, db)
        domain = _cafe24_store_domain(cafe24_user)

        # 1) 첫 번째 상품 정보로 대표 이미지/이름 — 표시용
        first_product_no = int(cafe24_product_nos[0])
        try:
            first_product = await cafe24_svc.get_product(cafe24_user, db, first_product_no)
        except Exception as e:
            logger.warning(f"[Campaign] 첫 상품 조회 실패: {e}")
            first_product = {}

        # 2) 비공개 카테고리 생성 — display=True여야 URL이 동작 (display=F면 카페24가 홈으로 302)
        #    use_main=False로 메인 메뉴엔 안 보이게 처리
        cat_name = cafe24_category_name or f"[비공개] {payload.name}"
        try:
            cat = await cafe24_svc.create_category(
                cafe24_user, db,
                category_name=cat_name,
                parent_category_no=cafe24_category_parent_no,
                display=True,
                use_main=False,
            )
            category_no = cat.get("category_no")
            if not category_no:
                raise RuntimeError("category_no 없음")
        except Exception as e:
            logger.error(f"[Campaign] 카테고리 생성 실패: {e}")
            raise HTTPException(
                status_code=502,
                detail=(
                    f"Cafe24 카테고리 생성 실패: {str(e)[:200]}. "
                    "OAuth scope에 mall.write_category 가 포함됐는지 확인 후 재연결하세요."
                ),
            )

        # 3) 상품들 카테고리에 attach
        try:
            attach_result = await cafe24_svc.attach_products_to_category(
                cafe24_user, db,
                category_no=int(category_no),
                product_nos=[int(p) for p in cafe24_product_nos],
            )
            logger.info(
                f"[Campaign] 카테고리 {category_no}에 상품 "
                f"{attach_result.get('attached', 0)}/{len(cafe24_product_nos)} 추가"
            )
        except Exception as e:
            logger.warning(f"[Campaign] 상품 attach 실패: {e}")
            category_warning = f"상품 일부 attach 실패: {str(e)[:200]}"

        # 4) 카테고리 URL 빌드
        category_url = cafe24_svc.category_storefront_url(domain, int(category_no))

        # DB 저장 필드
        data["cafe24_product_no"] = first_product_no  # 백워드 호환
        data["cafe24_product_name"] = first_product.get("product_name", "")
        data["cafe24_product_image"] = first_product.get("list_image", "")
        data["discount_type"] = discount_type
        data["discount_value"] = discount_value
        data["cafe24_category_no"] = int(category_no)
        data["cafe24_category_name"] = cat_name
        data["cafe24_category_url"] = category_url
        data["cafe24_product_nos"] = json.dumps([int(p) for p in cafe24_product_nos])
        data["base_product_url"] = category_url
        if not data.get("landing_url"):
            data["landing_url"] = category_url

        # 5) 카테고리 단위 쿠폰 발급 (할인이 설정된 경우만)
        if discount_type and discount_value is not None:
            try:
                benefit_type_map = {"percentage": "A", "fixed": "B", "shipping": "D"}
                benefit_type = benefit_type_map.get(discount_type, "A")
                coupon_result = await cafe24_svc.create_coupon(
                    cafe24_user, db,
                    coupon_name=f"[{payload.name}] 할인쿠폰",
                    benefit_type=benefit_type,
                    benefit_percentage=discount_value if discount_type == "percentage" else None,
                    benefit_price=discount_value if discount_type == "fixed" else None,
                    category_no=int(category_no),
                )
                data["cafe24_coupon_code"] = coupon_result.get("coupon_code")
                data["cafe24_coupon_no"] = coupon_result.get("coupon_no")
            except Exception as e:
                logger.warning(f"[Campaign] 카테고리 쿠폰 발급 실패: {e}")
                data["cafe24_coupon_code"] = None
                data["cafe24_coupon_no"] = None
                coupon_warning = f"쿠폰 발급 실패: {str(e)[:200]}"

    elif cafe24_product_no:
        # ─── A) 단일 상품 캠페인 (백워드 호환) ────────────────────────────────
        cafe24_user = await _resolve_cafe24_user(current_user, db)

        try:
            product = await cafe24_svc.get_product(cafe24_user, db, cafe24_product_no)
        except Exception as e:
            logger.warning(f"[Campaign] 상품 조회 실패: {e}")
            product = {}

        data["cafe24_product_no"] = cafe24_product_no
        data["cafe24_product_name"] = product.get("product_name", "")
        data["cafe24_product_image"] = product.get("list_image", "")
        data["discount_type"] = discount_type
        data["discount_value"] = discount_value

        domain = _cafe24_store_domain(cafe24_user)
        base_url = f"https://{domain}/product/detail.html?product_no={cafe24_product_no}"
        data["base_product_url"] = base_url

        accessible = await cafe24_svc.verify_storefront_url(base_url)
        if not accessible:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"선택한 상품(상품번호 {cafe24_product_no})이 스토어프론트에 공개되어 있지 않습니다. "
                    "Cafe24 관리자에서 진열/판매 상태, 카테고리 할당, 진열 사이트(PC/모바일)를 확인해주세요."
                ),
            )

        if not data.get("landing_url"):
            data["landing_url"] = base_url

        try:
            benefit_type_map = {"percentage": "A", "fixed": "B", "shipping": "D"}
            benefit_type = benefit_type_map.get(discount_type or "percentage", "A")
            coupon_result = await cafe24_svc.create_coupon(
                cafe24_user, db,
                coupon_name=f"[{payload.name}] 할인쿠폰",
                benefit_type=benefit_type,
                benefit_percentage=discount_value if discount_type == "percentage" else None,
                benefit_price=discount_value if discount_type == "fixed" else None,
                product_no=cafe24_product_no,
            )
            data["cafe24_coupon_code"] = coupon_result.get("coupon_code")
            data["cafe24_coupon_no"] = coupon_result.get("coupon_no")
        except Exception as e:
            logger.warning(f"[Campaign] 쿠폰 발급 실패: {e}")
            data["cafe24_coupon_code"] = None
            data["cafe24_coupon_no"] = None
            coupon_warning = f"쿠폰 발급 실패: {str(e)}"

    # 캠페인 자체 레퍼럴 코드 발급 (파트너 없이도 공유 가능)
    data["referral_code"] = await _unique_campaign_code(db)

    campaign = AffiliateCampaign(user_id=current_user.id, **data)
    db.add(campaign)
    await db.commit()
    await db.refresh(campaign)

    resp = {k: getattr(campaign, k) for k in campaign.__table__.columns.keys()}
    resp["referral_link"] = _build_referral_link(campaign, campaign.referral_code, current_user) if campaign.referral_code else None
    if coupon_warning:
        resp["warning"] = coupon_warning
    if category_warning:
        resp["category_warning"] = category_warning
    return resp


def _campaign_to_dict(c: AffiliateCampaign, user: Optional[User] = None) -> dict:
    """캠페인 ORM 객체 + 공유 링크 포함 dict."""
    d = {k: getattr(c, k) for k in c.__table__.columns.keys()}
    d["referral_link"] = _build_referral_link(c, c.referral_code, user) if c.referral_code else None
    return d


@router.get("/campaigns")
async def list_campaigns(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List all campaigns (전체 관리자 계정 공유) + 캠페인별 집계 포함."""
    result = await db.execute(select(AffiliateCampaign))
    campaigns = result.scalars().all()

    # 활성 파트너 ID 집합 (휴지통 제외) — 집계 시 필터
    active_pids_r = await db.execute(
        select(AffiliatePartner.id).where(AffiliatePartner.deleted_at.is_(None))
    )
    active_partner_ids = [row[0] for row in active_pids_r.all()]

    response = []
    for c in campaigns:
        d = _campaign_to_dict(c, current_user)
        pc_count_r = await db.execute(
            select(func.count(func.distinct(PartnerCampaign.partner_id))).where(
                PartnerCampaign.campaign_id == c.id,
                PartnerCampaign.partner_id.in_(active_partner_ids) if active_partner_ids else False,
            )
        )
        legacy_count_r = await db.execute(
            select(func.count(AffiliatePartner.id)).where(
                AffiliatePartner.campaign_id == c.id,
                AffiliatePartner.deleted_at.is_(None),
            )
        )
        click_conditions = [ReferralClick.campaign_id == c.id]
        conv_conditions = [
            ReferralConversion.campaign_id == c.id,
            ReferralConversion.status == "paid",
        ]
        if active_partner_ids:
            click_conditions.append(ReferralClick.partner_id.in_(active_partner_ids))
            conv_conditions.append(ReferralConversion.partner_id.in_(active_partner_ids))
        click_count_r = await db.execute(
            select(func.count(ReferralClick.id)).where(*click_conditions)
        )
        conv_r = await db.execute(
            select(
                func.count(ReferralConversion.id),
                func.coalesce(func.sum(ReferralConversion.order_amount), 0),
                func.coalesce(func.sum(ReferralConversion.commission_amount), 0),
            ).where(*conv_conditions)
        )
        conv_row = conv_r.one()
        clicks = click_count_r.scalar() or 0
        conversions = conv_row[0] or 0
        d["partner_count"] = (pc_count_r.scalar() or 0) + (legacy_count_r.scalar() or 0)
        d["click_count"] = clicks
        d["conversion_count"] = conversions
        d["total_sales"] = float(conv_row[1])
        d["total_commission"] = float(conv_row[2])
        d["conversion_rate"] = round((conversions / clicks * 100) if clicks > 0 else 0.0, 2)
        response.append(d)
    return response


@router.get("/campaigns/{campaign_id}")
async def get_campaign(
    campaign_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get a single campaign by ID (전체 관리자 공유)."""
    result = await db.execute(
        select(AffiliateCampaign).where(AffiliateCampaign.id == campaign_id)
    )
    campaign = result.scalar_one_or_none()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    return campaign


@router.put("/campaigns/{campaign_id}")
async def update_campaign(
    campaign_id: int,
    payload: CampaignUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Update an existing campaign (전체 관리자 공유)."""
    result = await db.execute(
        select(AffiliateCampaign).where(AffiliateCampaign.id == campaign_id)
    )
    campaign = result.scalar_one_or_none()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    for field, value in payload.model_dump(exclude_none=True).items():
        setattr(campaign, field, value)

    campaign.updated_at = datetime.utcnow()
    await db.commit()
    await db.refresh(campaign)
    return campaign


@router.delete("/campaigns/{campaign_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_campaign(
    campaign_id: int,
    skip_cafe24: bool = Query(False, description="카페24 카테고리 cleanup 건너뛰기 (강제 삭제)"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """캠페인 삭제 (전체 관리자 공유). 의존 레코드(partner_campaigns/clicks/conversions)도 함께 정리."""
    from sqlalchemy import update as sa_update

    result = await db.execute(
        select(AffiliateCampaign).where(AffiliateCampaign.id == campaign_id)
    )
    campaign = result.scalar_one_or_none()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    # 카페24 카테고리 cleanup (best-effort — 실패해도 캠페인 삭제는 진행)
    cafe24_category_no = campaign.cafe24_category_no
    if cafe24_category_no and not skip_cafe24:
        try:
            import app.services.cafe24 as cafe24_svc
            cafe24_user = await _resolve_cafe24_user(current_user, db)
            await cafe24_svc.delete_category(cafe24_user, db, category_no=int(cafe24_category_no))
            logger.info(f"[Affiliate] Cafe24 카테고리 {cafe24_category_no} 삭제 시도")
        except Exception as e:
            logger.warning(f"[Affiliate] Cafe24 카테고리 {cafe24_category_no} 삭제 실패 (무시하고 DB 삭제 진행): {e}")
    elif cafe24_category_no and skip_cafe24:
        logger.info(f"[Affiliate] Campaign {campaign_id}: 카페24 cleanup 건너뛰기 (skip_cafe24=true)")

    # FK 정리 + 본 삭제 — 에러 시 상세 메시지로 surface
    # 순서 중요: ReferralConversion.click_id → ReferralClick.id FK 때문에
    # click을 먼저 지우면 conversion이 그 click을 참조 중이라 FK 위반.
    # 따라서: conversions의 click_id NULL 처리 → conversions 삭제 → clicks 삭제 순서
    try:
        # 1) partner_campaigns 조인 테이블 행 삭제
        await db.execute(delete(PartnerCampaign).where(PartnerCampaign.campaign_id == campaign_id))
        # 2) affiliate_partners.campaign_id (레거시 단일 FK) NULL 처리
        await db.execute(
            sa_update(AffiliatePartner)
            .where(AffiliatePartner.campaign_id == campaign_id)
            .values(campaign_id=None)
        )
        # 3) 이 캠페인의 클릭을 참조하는 모든 conversion(다른 캠페인의 것 포함)의
        #    click_id를 NULL로 — 클릭 삭제 시 FK 위반 방지
        click_ids_subq = select(ReferralClick.id).where(ReferralClick.campaign_id == campaign_id).scalar_subquery()
        await db.execute(
            sa_update(ReferralConversion)
            .where(ReferralConversion.click_id.in_(click_ids_subq))
            .values(click_id=None)
        )
        # 4) 이 캠페인의 conversion 삭제 (집계에서 빠지도록)
        await db.execute(
            delete(ReferralConversion).where(ReferralConversion.campaign_id == campaign_id)
        )
        # 5) 이제 안전하게 click 삭제
        await db.execute(
            delete(ReferralClick).where(ReferralClick.campaign_id == campaign_id)
        )
        # 6) 캠페인 본체 삭제
        await db.delete(campaign)
        await db.commit()
        logger.info(f"[Affiliate] Campaign {campaign_id} deleted by user {current_user.id}")
    except Exception as e:
        await db.rollback()
        logger.exception(f"[Affiliate] Campaign {campaign_id} DB 삭제 실패")
        raise HTTPException(
            status_code=500,
            detail=(
                f"DB 삭제 실패: {type(e).__name__}: {str(e)[:300]}. "
                "외래키 제약 또는 다른 테이블 참조 가능성. 강제 삭제(skip_cafe24=true) 시도해보세요."
            ),
        )


# ---------------------------------------------------------------------------
# Partner CRUD
# ---------------------------------------------------------------------------

def _parse_channels(channels_str: Optional[str]) -> Optional[List[str]]:
    """JSON 문자열로 저장된 channels 컬럼을 list로 파싱."""
    if not channels_str:
        return None
    try:
        parsed = json.loads(channels_str)
        if isinstance(parsed, list):
            return parsed
    except (json.JSONDecodeError, ValueError):
        pass
    # comma-separated fallback
    return [c.strip() for c in channels_str.split(",") if c.strip()]


def _partner_to_dict(partner) -> dict:
    """AffiliatePartner ORM 객체를 dict로 변환하며 channels list 포함."""
    data = {k: getattr(partner, k) for k in partner.__table__.columns.keys()}
    data["channels"] = _parse_channels(partner.channels)
    return data


def _generate_referral_code() -> str:
    return uuid.uuid4().hex[:10].upper()


async def _unique_partner_code(db: AsyncSession) -> str:
    """충돌 없는 AffiliatePartner 레퍼럴 코드 생성."""
    for _ in range(10):
        code = _generate_referral_code()
        existing = await db.execute(
            select(AffiliatePartner).where(AffiliatePartner.referral_code == code)
        )
        if not existing.scalar_one_or_none():
            return code
    return _generate_referral_code()


async def _unique_pc_code(db: AsyncSession) -> str:
    """충돌 없는 PartnerCampaign 레퍼럴 코드 생성."""
    for _ in range(10):
        code = _generate_referral_code()
        existing = await db.execute(
            select(PartnerCampaign).where(PartnerCampaign.referral_code == code)
        )
        if not existing.scalar_one_or_none():
            return code
    return _generate_referral_code()


def _cafe24_store_domain(user: Optional[User]) -> str:
    """외부 노출용 Cafe24 스토어 도메인. CAFE24_PUBLIC_DOMAIN 우선, 없으면 mall_id.cafe24.com."""
    from app.core.config import get_settings as _gs
    _settings = _gs()
    if _settings.CAFE24_PUBLIC_DOMAIN:
        return _settings.CAFE24_PUBLIC_DOMAIN.replace("https://", "").replace("http://", "").rstrip("/")
    if user and user.cafe24_mall_id:
        return f"{user.cafe24_mall_id}.cafe24.com"
    return ""


def _build_destination_url(campaign: Optional[AffiliateCampaign], user: Optional[User]) -> Optional[str]:
    """
    Cafe24 실제 스토어프론트 도착 URL.

    우선순위:
      1) 비공개 카테고리 캠페인 (cafe24_category_no) → 카테고리 페이지 + 쿠폰 query
      2) 단일 상품 캠페인 (cafe24_product_no) → 상품 상세 + 쿠폰 query
      3) landing_url
      4) mall 도메인 루트

    카테고리 URL은 항상 최신 패턴(/product/list.html?cate_no=N)으로 재빌드 —
    DB의 cafe24_category_url에 옛 패턴이 캐시돼 있을 수 있어 무시.
    """
    if not campaign:
        return None
    domain = _cafe24_store_domain(user)

    # 1) 카테고리 캠페인 우선 — 항상 최신 패턴으로 재빌드
    if campaign.cafe24_category_no and domain:
        import app.services.cafe24 as _cafe24_svc
        url = _cafe24_svc.category_storefront_url(domain, int(campaign.cafe24_category_no))
        if campaign.cafe24_coupon_code:
            sep = "&" if "?" in url else "?"
            url = f"{url}{sep}coupon={campaign.cafe24_coupon_code}"
        return url

    # 2) 단일 상품 캠페인
    if campaign.cafe24_product_no and domain:
        url = f"https://{domain}/product/detail.html?product_no={campaign.cafe24_product_no}"
        if campaign.cafe24_coupon_code:
            url += f"&coupon={campaign.cafe24_coupon_code}"
        return url

    return campaign.landing_url or (f"https://{domain}" if domain else None)


def _build_referral_link(campaign: Optional[AffiliateCampaign], code: str, user: Optional[User] = None) -> str:
    """
    백엔드 추적기를 경유하는 레퍼럴 링크.
    `{BACKEND_URL}/r/{code}` — 클릭 시 ReferralClick 기록 후 Cafe24 상품 페이지로 302.
    Cafe24 직접 URL 대비 1 hop 느리지만 클릭/전환 추적이 가능.
    """
    from app.core.config import get_settings as _gs
    _settings = _gs()
    backend = (_settings.BACKEND_URL or "").rstrip("/")
    return f"{backend}/r/{code}" if backend else f"/r/{code}"


def _resolve_campaign_user_sync(campaign: AffiliateCampaign) -> Optional[User]:
    """캠페인 ORM 객체에서 user 조회 (동기 context에서는 None 반환 — 호출자가 user 전달 권장)."""
    return None


async def _unique_campaign_code(db: AsyncSession) -> str:
    """충돌 없는 AffiliateCampaign 레퍼럴 코드 생성."""
    for _ in range(10):
        code = _generate_referral_code()
        existing = await db.execute(
            select(AffiliateCampaign).where(AffiliateCampaign.referral_code == code)
        )
        if not existing.scalar_one_or_none():
            return code
    return _generate_referral_code()


@router.post("/partners", status_code=status.HTTP_201_CREATED)
async def create_partner(
    payload: PartnerCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Register a new affiliate partner and generate referral link."""
    code = await _unique_partner_code(db)

    # 첫 번째 캠페인 결정 (campaign_ids 우선)
    primary_campaign_id = payload.campaign_id
    if payload.campaign_ids:
        primary_campaign_id = payload.campaign_ids[0]

    # 기본 캠페인 조회
    campaign = None
    if primary_campaign_id:
        camp_result = await db.execute(
            select(AffiliateCampaign).where(AffiliateCampaign.id == primary_campaign_id)
        )
        campaign = camp_result.scalar_one_or_none()

    referral_link = _build_referral_link(campaign, code, current_user)

    partner_data = payload.model_dump(exclude={"campaign_ids", "channels"})
    partner_data["campaign_id"] = primary_campaign_id

    # channels 다중 선택 처리
    if payload.channels:
        partner_data["channels"] = json.dumps(payload.channels)
        # 기존 channel 컬럼은 channels[0] 으로 채움 (하위 호환)
        partner_data["channel"] = payload.channels[0]
    # channel은 payload에서 이미 설정됨 (default "instagram")

    # partner_group이 None이면 모델 default("crew")를 사용하도록 키 제거
    if partner_data.get("partner_group") is None:
        partner_data.pop("partner_group", None)

    # 관리자가 직접 초대한 파트너는 즉시 승인 — SMS 매직링크가 바로 동작해야 함
    partner_data.setdefault("status", "approved")
    partner = AffiliatePartner(
        user_id=current_user.id,
        referral_code=code,
        referral_link=referral_link,
        **partner_data,
    )
    db.add(partner)
    await db.flush()  # partner.id 확보

    # Phase 5: campaign_ids 각각에 대해 PartnerCampaign 생성
    all_campaign_ids = list(dict.fromkeys(
        ([primary_campaign_id] if primary_campaign_id else []) + payload.campaign_ids
    ))
    for cid in all_campaign_ids:
        camp_result = await db.execute(
            select(AffiliateCampaign).where(AffiliateCampaign.id == cid)
        )
        c = camp_result.scalar_one_or_none()
        if not c:
            continue
        pc_code = await _unique_pc_code(db)
        pc_link = _build_referral_link(c, pc_code, current_user)
        pc = PartnerCampaign(
            partner_id=partner.id,
            campaign_id=cid,
            referral_code=pc_code,
            referral_link=pc_link,
        )
        db.add(pc)

    await db.commit()
    await db.refresh(partner)

    # 이메일 발송 (Resend)
    if partner.email:
        try:
            if _settings.RESEND_API_KEY:
                import httpx
                async with httpx.AsyncClient(timeout=10) as hclient:
                    await hclient.post(
                        "https://api.resend.com/emails",
                        headers={"Authorization": f"Bearer {_settings.RESEND_API_KEY}"},
                        json={
                            "from": _settings.RESEND_FROM_EMAIL or "noreply@joinandjoin.com",
                            "to": [partner.email],
                            "subject": "[널담] 어필리에이트 파트너 초대",
                            "html": (
                                f"<h2>안녕하세요, {partner.name}님!</h2>"
                                f"<p>널담 어필리에이트 파트너로 초대되었습니다.</p>"
                                f"<p>아래 전용 링크를 통해 상품을 홍보하고 커미션을 받으세요:</p>"
                                f"<p><a href='{referral_link}' style='font-size:18px;font-weight:bold;'>{referral_link}</a></p>"
                                f"<p>레퍼럴 코드: <b>{code}</b></p>"
                                f"<br><p>감사합니다,<br>널담은디저트</p>"
                            ),
                        },
                    )
                logger.info(f"[Affiliate] Invite email sent to {partner.email}")
        except Exception as e:
            logger.warning(f"[Affiliate] Email send failed: {e}")

    # SMS 발송 (Solapi) — phone 있을 때, 자동 로그인 매직링크 포함
    if partner.phone:
        try:
            from app.services.sms import send_sms
            from app.core.security import create_access_token
            from datetime import timedelta as _td
            # 30일 유효 매직링크 토큰 — 클릭 시 즉시 로그인되어 별도 입력 불필요
            invite_token = create_access_token(
                data={"sub": str(partner.id), "type": "partner_magic_link"},
                expires_delta=_td(days=30),
            )
            portal_url = f"{_settings.FRONTEND_URL}/partner?token={invite_token}"
            sms_message = (
                f"[널담] {partner.name}님, 어필리에이트 파트너로 초대되었습니다.\n"
                f"아래 링크를 누르면 바로 판매 현황 페이지로 이동합니다.\n{portal_url}"
            )
            sms_result = await send_sms(partner.phone, sms_message)
            if sms_result.get("success"):
                logger.info(f"[Affiliate] Invite SMS sent to {partner.phone}")
            else:
                logger.warning(f"[Affiliate] SMS send skipped/failed: {sms_result.get('reason')}")
        except Exception as e:
            logger.warning(f"[Affiliate] SMS send failed: {e}")

    return _partner_to_dict(partner)


@router.get("/partners")
async def list_partners(
    campaign_id: Optional[int] = None,
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List partners + 파트너별 집계 필드 포함."""
    query = select(AffiliatePartner).where(AffiliatePartner.deleted_at.is_(None))
    if campaign_id is not None:
        query = query.where(AffiliatePartner.campaign_id == campaign_id)
    if status is not None:
        query = query.where(AffiliatePartner.status == status)

    result = await db.execute(query)
    partners = result.scalars().all()

    response = []
    for p in partners:
        d = _partner_to_dict(p)
        click_r = await db.execute(
            select(func.count(ReferralClick.id)).where(ReferralClick.partner_id == p.id)
        )
        conv_r = await db.execute(
            select(
                func.count(ReferralConversion.id),
                func.coalesce(func.sum(ReferralConversion.order_amount), 0),
                func.coalesce(func.sum(ReferralConversion.commission_amount), 0),
            ).where(
                ReferralConversion.partner_id == p.id,
                ReferralConversion.status == "paid",
            )
        )
        conv_row = conv_r.one()
        settle_r = await db.execute(
            select(func.coalesce(func.sum(AffiliateSettlement.amount), 0)).where(
                AffiliateSettlement.partner_id == p.id,
                AffiliateSettlement.status == "paid",
            )
        )
        clicks = click_r.scalar() or 0
        conversions = conv_row[0] or 0
        total_commission = float(conv_row[2])
        paid_settlement = float(settle_r.scalar() or 0)
        d["click_count"] = clicks
        d["conversion_count"] = conversions
        d["total_sales"] = float(conv_row[1])
        d["total_commission"] = total_commission
        d["unpaid_commission"] = max(0.0, total_commission - paid_settlement)
        d["conversion_rate"] = round((conversions / clicks * 100) if clicks > 0 else 0.0, 2)

        # 파트너별 연결된 캠페인들의 링크 배열 (중복 제거)
        pc_rows_r = await db.execute(
            select(PartnerCampaign).where(PartnerCampaign.partner_id == p.id)
        )
        pcs = list(pc_rows_r.scalars().all())
        seen_campaign_ids = {pc.campaign_id for pc in pcs}
        campaign_links = []
        for pc in pcs:
            cr = await db.execute(
                select(AffiliateCampaign).where(AffiliateCampaign.id == pc.campaign_id)
            )
            camp = cr.scalar_one_or_none()
            campaign_links.append({
                "pc_id": pc.id,
                "campaign_id": pc.campaign_id,
                "campaign_name": camp.name if camp else "",
                "referral_code": pc.referral_code,
                "referral_link": _build_referral_link(camp, pc.referral_code, current_user) if camp else pc.referral_link,
            })
        # legacy campaign_id 가상 row (PC에 없고 partner.campaign_id만 있는 경우)
        if p.campaign_id and p.campaign_id not in seen_campaign_ids:
            cr = await db.execute(
                select(AffiliateCampaign).where(AffiliateCampaign.id == p.campaign_id)
            )
            camp = cr.scalar_one_or_none()
            campaign_links.append({
                "pc_id": -1,
                "campaign_id": p.campaign_id,
                "campaign_name": camp.name if camp else "",
                "referral_code": p.referral_code,
                "referral_link": _build_referral_link(camp, p.referral_code, current_user) if camp else p.referral_link,
            })
        d["campaign_links"] = campaign_links
        response.append(d)
    return response


@router.put("/partners/{partner_id}")
async def update_partner(
    partner_id: int,
    payload: PartnerUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Update partner information (전체 관리자 공유)."""
    result = await db.execute(
        select(AffiliatePartner).where(AffiliatePartner.id == partner_id)
    )
    partner = result.scalar_one_or_none()
    if not partner:
        raise HTTPException(status_code=404, detail="Partner not found")

    update_data = payload.model_dump(exclude_none=True, exclude={"channels"})
    for field, value in update_data.items():
        setattr(partner, field, value)

    # channels 다중 선택 처리
    if payload.channels is not None:
        partner.channels = json.dumps(payload.channels)
        # 기존 channel 컬럼도 갱신 (하위 호환)
        partner.channel = payload.channels[0] if payload.channels else partner.channel

    await db.commit()
    await db.refresh(partner)
    return _partner_to_dict(partner)


@router.post("/partners/{partner_id}/approve")
async def approve_partner(
    partner_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Approve a pending partner (전체 관리자 공유)."""
    result = await db.execute(
        select(AffiliatePartner).where(AffiliatePartner.id == partner_id)
    )
    partner = result.scalar_one_or_none()
    if not partner:
        raise HTTPException(status_code=404, detail="Partner not found")

    partner.status = "approved"
    await db.commit()
    await db.refresh(partner)
    return {"success": True, "partner_id": partner_id, "status": partner.status}


@router.post("/partners/{partner_id}/reject")
async def reject_partner(
    partner_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Reject a pending partner (전체 관리자 공유)."""
    result = await db.execute(
        select(AffiliatePartner).where(AffiliatePartner.id == partner_id)
    )
    partner = result.scalar_one_or_none()
    if not partner:
        raise HTTPException(status_code=404, detail="Partner not found")

    partner.status = "rejected"
    await db.commit()
    await db.refresh(partner)
    return {"success": True, "partner_id": partner_id, "status": partner.status}


@router.delete("/partners/{partner_id}", status_code=status.HTTP_200_OK)
async def delete_partner(
    partner_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """파트너 소프트 삭제 — 휴지통으로 이동 (전체 관리자 공유)."""
    result = await db.execute(
        select(AffiliatePartner).where(AffiliatePartner.id == partner_id)
    )
    partner = result.scalar_one_or_none()
    if not partner:
        raise HTTPException(status_code=404, detail="Partner not found")

    partner.deleted_at = datetime.utcnow()
    await db.commit()
    logger.info(f"[Affiliate] Partner {partner_id} moved to trash by user {current_user.id}")
    return {"success": True, "partner_id": partner_id, "trashed_at": partner.deleted_at}


@router.get("/partners/trash")
async def list_deleted_partners(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """휴지통에 있는 파트너 목록 (전체 관리자 공유)."""
    result = await db.execute(
        select(AffiliatePartner).where(
            AffiliatePartner.deleted_at.isnot(None),
        ).order_by(AffiliatePartner.deleted_at.desc())
    )
    return [_partner_to_dict(p) for p in result.scalars().all()]


@router.post("/partners/{partner_id}/restore")
async def restore_partner(
    partner_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """휴지통에서 파트너 복원 (전체 관리자 공유)."""
    result = await db.execute(
        select(AffiliatePartner).where(AffiliatePartner.id == partner_id)
    )
    partner = result.scalar_one_or_none()
    if not partner:
        raise HTTPException(status_code=404, detail="Partner not found")
    if partner.deleted_at is None:
        raise HTTPException(status_code=400, detail="이미 활성 상태인 파트너입니다.")

    partner.deleted_at = None
    await db.commit()
    return {"success": True, "partner_id": partner_id, "status": partner.status}


@router.delete("/partners/{partner_id}/permanent", status_code=status.HTTP_204_NO_CONTENT)
async def permanent_delete_partner(
    partner_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """영구 삭제 (전체 관리자 공유) — PartnerCampaign/Click/Conversion/Settlement 전부 삭제 후 파트너 제거."""
    result = await db.execute(
        select(AffiliatePartner).where(AffiliatePartner.id == partner_id)
    )
    partner = result.scalar_one_or_none()
    if not partner:
        raise HTTPException(status_code=404, detail="Partner not found")

    await db.execute(delete(PartnerCampaign).where(PartnerCampaign.partner_id == partner_id))
    await db.execute(delete(ReferralClick).where(ReferralClick.partner_id == partner_id))
    await db.execute(delete(ReferralConversion).where(ReferralConversion.partner_id == partner_id))
    await db.execute(delete(AffiliateSettlement).where(AffiliateSettlement.partner_id == partner_id))
    await db.delete(partner)
    await db.commit()
    logger.info(f"[Affiliate] Partner {partner_id} permanently deleted by user {current_user.id}")


# ---------------------------------------------------------------------------
# Dashboard stats
# ---------------------------------------------------------------------------

@router.get("/dashboard")
async def get_dashboard(
    days: Optional[int] = Query(default=None, ge=1, le=365),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Return aggregated affiliate KPIs for the current user.

    days 지정 시 클릭/전환/매출 집계를 최근 N일로 제한 (파트너 수·정산 대기 등 비시계열 지표는 전체 기준 유지).
    """
    since = datetime.utcnow() - timedelta(days=days) if days else None

    # Total partners by status (휴지통 제외, 전체 공유)
    partners_result = await db.execute(
        select(AffiliatePartner.status, func.count(AffiliatePartner.id))
        .where(AffiliatePartner.deleted_at.is_(None))
        .group_by(AffiliatePartner.status)
    )
    partners_by_status = {row[0]: row[1] for row in partners_result.all()}

    # 활성 파트너 ID만 (휴지통 파트너 제외)
    partner_ids_result = await db.execute(
        select(AffiliatePartner.id).where(AffiliatePartner.deleted_at.is_(None))
    )
    partner_ids = [row[0] for row in partner_ids_result.all()]

    # Total clicks
    total_clicks = 0
    if partner_ids:
        click_conds = [ReferralClick.partner_id.in_(partner_ids)]
        if since is not None:
            click_conds.append(ReferralClick.clicked_at >= since)
        clicks_result = await db.execute(
            select(func.count(ReferralClick.id)).where(*click_conds)
        )
        total_clicks = clicks_result.scalar() or 0

    # Total conversions + revenue + commission (status='paid' 순매출만 집계)
    total_conversions = 0
    total_revenue = 0.0
    total_commission = 0.0
    refunded_count = 0
    cancelled_count = 0
    gross_sales = 0.0
    if partner_ids:
        # 순매출 (paid only)
        paid_conds = [
            ReferralConversion.partner_id.in_(partner_ids),
            ReferralConversion.status == "paid",
        ]
        if since is not None:
            paid_conds.append(ReferralConversion.converted_at >= since)
        conv_result = await db.execute(
            select(
                func.count(ReferralConversion.id),
                func.coalesce(func.sum(ReferralConversion.order_amount), 0),
                func.coalesce(func.sum(ReferralConversion.commission_amount), 0),
            ).where(*paid_conds)
        )
        row = conv_result.one()
        total_conversions = row[0] or 0
        total_revenue = float(row[1])
        total_commission = float(row[2])

        # 환불/취소 건수
        rc_conds = [
            ReferralConversion.partner_id.in_(partner_ids),
            ReferralConversion.status.in_(["refunded", "cancelled"]),
        ]
        if since is not None:
            rc_conds.append(ReferralConversion.converted_at >= since)
        status_result = await db.execute(
            select(
                ReferralConversion.status,
                func.count(ReferralConversion.id),
            ).where(*rc_conds).group_by(ReferralConversion.status)
        )
        for srow in status_result.all():
            if srow[0] == "refunded":
                refunded_count = srow[1]
            elif srow[0] == "cancelled":
                cancelled_count = srow[1]

        # 총 gross (모든 상태 합, 참고용)
        gross_conds = [ReferralConversion.partner_id.in_(partner_ids)]
        if since is not None:
            gross_conds.append(ReferralConversion.converted_at >= since)
        gross_result = await db.execute(
            select(func.coalesce(func.sum(ReferralConversion.order_amount), 0)).where(*gross_conds)
        )
        gross_sales = float(gross_result.scalar() or 0)

    # Conversion rate
    conversion_rate = (total_conversions / total_clicks * 100) if total_clicks > 0 else 0.0

    # Total campaigns (전체 공유)
    campaigns_result = await db.execute(
        select(func.count(AffiliateCampaign.id))
    )
    total_campaigns = campaigns_result.scalar() or 0

    # Pending settlements (전체 공유)
    pending_settlement_result = await db.execute(
        select(func.coalesce(func.sum(AffiliateSettlement.amount), 0)).where(
            AffiliateSettlement.status == "pending",
        )
    )
    pending_settlement_amount = float(pending_settlement_result.scalar() or 0)

    # Active campaigns (전체 공유, 각 캠페인별 집계 포함)
    active_campaigns_result = await db.execute(
        select(AffiliateCampaign).where(
            AffiliateCampaign.status == "active",
        ).limit(5)
    )
    active_campaigns = []
    for c in active_campaigns_result.scalars().all():
        # 캠페인별 파트너 수 (PartnerCampaign join + legacy campaign_id)
        pc_pcount = await db.execute(
            select(func.count(func.distinct(PartnerCampaign.partner_id))).where(
                PartnerCampaign.campaign_id == c.id
            )
        )
        pc_count = pc_pcount.scalar() or 0
        legacy_pcount = await db.execute(
            select(func.count(AffiliatePartner.id)).where(
                AffiliatePartner.campaign_id == c.id,
                AffiliatePartner.deleted_at.is_(None),
            )
        )
        partner_count = (pc_count or 0) + (legacy_pcount.scalar() or 0)

        # 클릭/전환/매출 (전환은 paid만)
        camp_click_conds = [ReferralClick.campaign_id == c.id]
        camp_conv_conds = [
            ReferralConversion.campaign_id == c.id,
            ReferralConversion.status == "paid",
        ]
        if since is not None:
            camp_click_conds.append(ReferralClick.clicked_at >= since)
            camp_conv_conds.append(ReferralConversion.converted_at >= since)
        click_count_r = await db.execute(
            select(func.count(ReferralClick.id)).where(*camp_click_conds)
        )
        conv_r = await db.execute(
            select(
                func.count(ReferralConversion.id),
                func.coalesce(func.sum(ReferralConversion.order_amount), 0),
                func.coalesce(func.sum(ReferralConversion.commission_amount), 0),
            ).where(*camp_conv_conds)
        )
        conv_row = conv_r.one()

        active_campaigns.append({
            "id": c.id,
            "name": c.name,
            "product": c.product,
            "commission_type": c.commission_type,
            "commission_rate": c.commission_rate,
            "status": c.status,
            "partner_count": partner_count,
            "click_count": click_count_r.scalar() or 0,
            "conversion_count": conv_row[0] or 0,
            "total_sales": float(conv_row[1]),
            "total_commission": float(conv_row[2]),
        })

    # Top partners by conversion count (전체 공유, 휴지통 제외)
    top_partners = []
    top_join_cond = (
        (ReferralConversion.partner_id == AffiliatePartner.id)
        & (ReferralConversion.status == "paid")
    )
    if since is not None:
        top_join_cond = top_join_cond & (ReferralConversion.converted_at >= since)
    top_result = await db.execute(
        select(
            AffiliatePartner.id, AffiliatePartner.name, AffiliatePartner.channel,
            AffiliatePartner.channels,
            AffiliatePartner.followers,
            func.count(ReferralConversion.id).label("conversions"),
            func.coalesce(func.sum(ReferralConversion.order_amount), 0).label("sales"),
        )
        .outerjoin(ReferralConversion, top_join_cond)
        .where(
            AffiliatePartner.status == "approved",
            AffiliatePartner.deleted_at.is_(None),
        )
        .group_by(AffiliatePartner.id)
        .order_by(func.coalesce(func.sum(ReferralConversion.order_amount), 0).desc())
        .limit(5)
    )
    for row in top_result.all():
        top_partners.append({
            "id": row[0], "name": row[1], "channel": row[2],
            "channels": _parse_channels(row[3]),
            "followers": row[4],
            "conversion_count": row[5], "total_sales": float(row[6]),
        })

    return {
        "total_campaigns": total_campaigns,
        "total_sales": total_revenue,       # status='paid' 순매출
        "net_sales": total_revenue,         # 명시적 순매출 alias
        "gross_sales": gross_sales,         # 모든 상태 합 (참고용)
        "total_commission": total_commission,
        "active_partners": partners_by_status.get("approved", 0),
        "total_partners": sum(partners_by_status.values()),
        "partners_by_status": partners_by_status,
        "total_clicks": total_clicks,
        "total_conversions": total_conversions,
        "refunded_count": refunded_count,
        "cancelled_count": cancelled_count,
        "conversion_rate": round(conversion_rate, 2),
        "pending_settlement_amount": pending_settlement_amount,
        "active_campaigns": active_campaigns,
        "top_partners": top_partners,
    }


@router.get("/dashboard/timeseries")
async def get_dashboard_timeseries(
    days: int = Query(default=30, ge=1, le=365),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    최근 N일간 일별 매출/커미션/클릭/전환 시계열 데이터 (전체 관리자 공유).

    데이터가 있는 날짜만 반환하며 날짜 gap은 프론트엔드에서 채웁니다.
    """
    # 활성 파트너 ID만 (휴지통 제외)
    pid_result = await db.execute(
        select(AffiliatePartner.id).where(AffiliatePartner.deleted_at.is_(None))
    )
    partner_ids = [row[0] for row in pid_result.all()]

    since = datetime.utcnow() - timedelta(days=days)

    # 일별 클릭 집계 — ORM으로 작성 (asyncpg의 ANY 바인딩 이슈 회피)
    clicks_by_date: dict = {}
    if partner_ids:
        day_col = func.date(ReferralClick.clicked_at).label("day")
        click_rows = await db.execute(
            select(day_col, func.count(ReferralClick.id))
            .where(
                ReferralClick.partner_id.in_(partner_ids),
                ReferralClick.clicked_at >= since,
            )
            .group_by(day_col)
            .order_by(day_col)
        )
        for row in click_rows.all():
            clicks_by_date[str(row[0])] = int(row[1])

    # 일별 전환/매출/커미션 집계 (status별 — paid/refunded/cancelled 분리)
    conv_by_date: dict = {}
    if partner_ids:
        day_col2 = func.date(ReferralConversion.converted_at).label("day")
        conv_rows = await db.execute(
            select(
                day_col2,
                ReferralConversion.status,
                func.count(ReferralConversion.id),
                func.coalesce(func.sum(ReferralConversion.order_amount), 0),
                func.coalesce(func.sum(ReferralConversion.commission_amount), 0),
            )
            .where(
                ReferralConversion.partner_id.in_(partner_ids),
                ReferralConversion.converted_at >= since,
            )
            .group_by(day_col2, ReferralConversion.status)
            .order_by(day_col2)
        )
        for row in conv_rows.all():
            date_str = str(row[0])
            st = row[1] or "paid"
            entry = conv_by_date.setdefault(date_str, {
                "conversions": 0, "revenue": 0.0, "commission": 0.0,
                "refunded_count": 0, "refunded_amount": 0.0,
                "cancelled_count": 0, "cancelled_amount": 0.0,
            })
            count = int(row[2])
            amount = float(row[3])
            commission = float(row[4])
            if st == "paid":
                entry["conversions"] = count
                entry["revenue"] = amount
                entry["commission"] = commission
            elif st == "refunded":
                entry["refunded_count"] = count
                entry["refunded_amount"] = amount
            elif st == "cancelled":
                entry["cancelled_count"] = count
                entry["cancelled_amount"] = amount

    # 날짜 합치기
    all_dates = sorted(set(list(clicks_by_date.keys()) + list(conv_by_date.keys())))
    result_list = []
    for date_str in all_dates:
        conv = conv_by_date.get(date_str, {
            "conversions": 0, "revenue": 0.0, "commission": 0.0,
            "refunded_count": 0, "refunded_amount": 0.0,
            "cancelled_count": 0, "cancelled_amount": 0.0,
        })
        result_list.append({
            "date": date_str,
            "revenue": conv["revenue"],
            "commission": conv["commission"],
            "clicks": clicks_by_date.get(date_str, 0),
            "conversions": conv["conversions"],
            "refunded_count": conv.get("refunded_count", 0),
            "refunded_amount": conv.get("refunded_amount", 0.0),
            "cancelled_count": conv.get("cancelled_count", 0),
            "cancelled_amount": conv.get("cancelled_amount", 0.0),
        })

    return result_list


@router.get("/dashboard/by-campaign")
async def get_dashboard_by_campaign(
    days: Optional[int] = Query(default=None, ge=1, le=365),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    캠페인별 성과 집계: 매출/커미션/클릭/전환/파트너 수 (전체 관리자 공유).
    days 지정 시 클릭/전환 집계를 최근 N일로 제한.
    """
    since = datetime.utcnow() - timedelta(days=days) if days else None
    # 활성 파트너 ID (휴지통 제외)
    active_pids_r = await db.execute(
        select(AffiliatePartner.id).where(AffiliatePartner.deleted_at.is_(None))
    )
    active_partner_ids = [row[0] for row in active_pids_r.all()]

    # 전체 캠페인 목록
    camp_result = await db.execute(select(AffiliateCampaign))
    campaigns = camp_result.scalars().all()

    result_list = []
    for campaign in campaigns:
        pcount_result = await db.execute(
            select(func.count(AffiliatePartner.id)).where(
                AffiliatePartner.campaign_id == campaign.id,
                AffiliatePartner.deleted_at.is_(None),
            )
        )
        partner_count = pcount_result.scalar() or 0

        click_conds = [ReferralClick.campaign_id == campaign.id]
        conv_conds = [
            ReferralConversion.campaign_id == campaign.id,
            ReferralConversion.status == "paid",
        ]
        if since is not None:
            click_conds.append(ReferralClick.clicked_at >= since)
            conv_conds.append(ReferralConversion.converted_at >= since)
        if active_partner_ids:
            click_conds.append(ReferralClick.partner_id.in_(active_partner_ids))
            conv_conds.append(ReferralConversion.partner_id.in_(active_partner_ids))

        click_result = await db.execute(select(func.count(ReferralClick.id)).where(*click_conds))
        clicks = click_result.scalar() or 0

        conv_result = await db.execute(
            select(
                func.count(ReferralConversion.id),
                func.coalesce(func.sum(ReferralConversion.order_amount), 0),
                func.coalesce(func.sum(ReferralConversion.commission_amount), 0),
            ).where(*conv_conds)
        )
        conv_row = conv_result.one()
        conversions = conv_row[0] or 0
        revenue = float(conv_row[1])
        commission = float(conv_row[2])

        result_list.append({
            "campaign_id": campaign.id,
            "campaign_name": campaign.name,
            "revenue": revenue,
            "commission": commission,
            "clicks": clicks,
            "conversions": conversions,
            "partners": partner_count,
        })

    return result_list


# ---------------------------------------------------------------------------
# Hourly heatmap — 시간대 x 요일별 전환 매트릭스
# ---------------------------------------------------------------------------

@router.get("/dashboard/hourly")
async def get_dashboard_hourly(
    days: int = Query(default=30, ge=1, le=365),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    최근 N일 내 status='paid' 전환을 시간대(0-23) x 요일(0=월, 6=일) 매트릭스로 집계.

    응답: 168 rows (7일 * 24시간). 데이터 없는 셀은 0으로 채워 반환.
    PostgreSQL EXTRACT(ISODOW): 1=월 ~ 7=일 → 0-based로 변환 (ISODOW - 1).
    """
    from sqlalchemy import text as sa_text

    # 빈 매트릭스 초기화 (0=월 ~ 6=일, 0~23시)
    matrix: dict[tuple[int, int], dict] = {}
    for dow in range(7):
        for hour in range(24):
            matrix[(hour, dow)] = {"conversions": 0, "revenue": 0.0}

    since = datetime.utcnow() - timedelta(days=days)

    # 활성 파트너만 (휴지통 제외)
    active_pids_r = await db.execute(
        select(AffiliatePartner.id).where(AffiliatePartner.deleted_at.is_(None))
    )
    active_partner_ids = [row[0] for row in active_pids_r.all()]

    # EXTRACT(ISODOW) - 1 → 0=월, 6=일
    hour_col = func.extract("hour", ReferralConversion.converted_at).label("hour")
    dow_col = (func.extract("isodow", ReferralConversion.converted_at) - 1).label("dow")

    conds = [
        ReferralConversion.converted_at >= since,
        ReferralConversion.status == "paid",
    ]
    if active_partner_ids:
        conds.append(ReferralConversion.partner_id.in_(active_partner_ids))
    else:
        return [
            {"hour": h, "day_of_week": d, "conversions": 0, "revenue": 0.0}
            for d in range(7) for h in range(24)
        ]

    rows = await db.execute(
        select(
            hour_col,
            dow_col,
            func.count(ReferralConversion.id),
            func.coalesce(func.sum(ReferralConversion.order_amount), 0),
        )
        .where(*conds)
        .group_by(hour_col, dow_col)
    )

    for row in rows.all():
        h = int(row[0])
        d = int(row[1])
        matrix[(h, d)] = {
            "conversions": int(row[2]),
            "revenue": float(row[3]),
        }

    result_list = [
        {
            "hour": h,
            "day_of_week": d,
            "conversions": matrix[(h, d)]["conversions"],
            "revenue": matrix[(h, d)]["revenue"],
        }
        for d in range(7) for h in range(24)
    ]
    logger.info(
        f"[Dashboard] hourly matrix: user={current_user.id} days={days} (shared/all partners)"
    )
    return result_list


# ---------------------------------------------------------------------------
# Top-products — 상품별 성과 TOP N
# ---------------------------------------------------------------------------

@router.get("/dashboard/top-products")
async def get_dashboard_top_products(
    limit: int = Query(default=10, ge=1, le=100),
    days: Optional[int] = Query(default=None, ge=1, le=365),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    캠페인의 cafe24_product_no 기준으로 그룹핑한 상품별 성과.

    status='paid' 전환만 포함. revenue 내림차순 정렬. days 지정 시 최근 N일만.
    """
    since = datetime.utcnow() - timedelta(days=days) if days else None
    # 활성 파트너만 (휴지통 제외)
    active_pids_r = await db.execute(
        select(AffiliatePartner.id).where(AffiliatePartner.deleted_at.is_(None))
    )
    active_partner_ids = [row[0] for row in active_pids_r.all()]

    # 전체 캠페인 중 cafe24_product_no가 있는 것만 (전체 관리자 공유)
    # GROUP BY product_no + name + image
    join_cond = (
        (ReferralConversion.campaign_id == AffiliateCampaign.id)
        & (ReferralConversion.status == "paid")
    )
    if since is not None:
        join_cond = join_cond & (ReferralConversion.converted_at >= since)
    if active_partner_ids:
        join_cond = join_cond & ReferralConversion.partner_id.in_(active_partner_ids)
    else:
        # 활성 파트너 없음 → 전환은 모두 0으로 반환
        join_cond = join_cond & (ReferralConversion.id == None)  # noqa: E711

    rows = await db.execute(
        select(
            AffiliateCampaign.cafe24_product_no,
            AffiliateCampaign.cafe24_product_name,
            AffiliateCampaign.cafe24_product_image,
            func.count(func.distinct(AffiliateCampaign.id)).label("campaign_count"),
            func.count(ReferralConversion.id).label("conversions"),
            func.coalesce(func.sum(ReferralConversion.order_amount), 0).label("revenue"),
            func.coalesce(func.sum(ReferralConversion.commission_amount), 0).label("commission"),
        )
        .outerjoin(ReferralConversion, join_cond)
        .where(
            AffiliateCampaign.cafe24_product_no.isnot(None),
        )
        .group_by(
            AffiliateCampaign.cafe24_product_no,
            AffiliateCampaign.cafe24_product_name,
            AffiliateCampaign.cafe24_product_image,
        )
        .order_by(func.coalesce(func.sum(ReferralConversion.order_amount), 0).desc())
        .limit(limit)
    )

    result_list = []
    for row in rows.all():
        result_list.append({
            "product_no": row[0],
            "product_name": row[1] or "",
            "product_image": row[2],
            "campaign_count": int(row[3]),
            "conversions": int(row[4]),
            "revenue": float(row[5]),
            "commission": float(row[6]),
        })

    logger.info(
        f"[Dashboard] top-products: user={current_user.id} limit={limit} results={len(result_list)}"
    )
    return result_list


# ---------------------------------------------------------------------------
# Settlements
# ---------------------------------------------------------------------------

@router.get("/settlements")
async def list_settlements(
    partner_id: Optional[int] = None,
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List settlements (전체 관리자 공유)."""
    query = select(AffiliateSettlement)
    if partner_id is not None:
        query = query.where(AffiliateSettlement.partner_id == partner_id)
    if status is not None:
        query = query.where(AffiliateSettlement.status == status)

    result = await db.execute(query)
    return result.scalars().all()


@router.post("/settlements", status_code=status.HTTP_201_CREATED)
async def create_settlement(
    payload: SettlementCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Create a settlement record for a partner."""
    # 파트너 존재 확인 (전체 공유)
    partner_result = await db.execute(
        select(AffiliatePartner).where(AffiliatePartner.id == payload.partner_id)
    )
    if not partner_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Partner not found")

    settlement = AffiliateSettlement(
        user_id=current_user.id,
        partner_id=payload.partner_id,
        amount=payload.amount,
        period_start=payload.period_start,
        period_end=payload.period_end,
        status="pending",
    )
    db.add(settlement)
    await db.commit()
    await db.refresh(settlement)
    return settlement


@router.post("/settlements/{settlement_id}/pay")
async def pay_settlement(
    settlement_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Mark a settlement as paid (전체 관리자 공유)."""
    result = await db.execute(
        select(AffiliateSettlement).where(AffiliateSettlement.id == settlement_id)
    )
    settlement = result.scalar_one_or_none()
    if not settlement:
        raise HTTPException(status_code=404, detail="Settlement not found")
    if settlement.status == "paid":
        raise HTTPException(status_code=400, detail="Settlement is already paid")

    settlement.status = "paid"
    settlement.paid_at = datetime.utcnow()
    await db.commit()
    await db.refresh(settlement)
    return {"success": True, "settlement_id": settlement_id, "paid_at": settlement.paid_at}


# ---------------------------------------------------------------------------
# Settlement Excel Export — 파트너별 [전체주문건][취소건] 2탭 XLSX
# ---------------------------------------------------------------------------

@router.get("/partners/{partner_id}/settlement-export")
async def export_partner_settlement(
    partner_id: int,
    start: Optional[str] = Query(None, description="YYYY-MM-DD"),
    end: Optional[str] = Query(None, description="YYYY-MM-DD"),
    seller_type: str = Query("freelancer", description="freelancer | business"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    파트너별 정산서 엑셀 다운로드 (3시트 구성).

    Sheet1 '요약': 판매자 유형(프리랜서/사업자)에 따른 정산서 양식
      - 프리랜서: 총공급가(부가세 별도) + 소득세 3% + 주민세 0.3% 차감
      - 사업자: 부가세 별도 처리, 세금 차감 없음 (본인 세금계산서 발행)
    Sheet2 '전체주문건': 모든 전환(주문) — 상품명/수량/수수료 컬럼 + SUM 수식
    Sheet3 '취소건': status in (refunded, cancelled) 또는 부분 환불

    검토 프로세스: 정상주문 합계 - 취소/환불 합계 = 유효 주문 → 수수료율 적용
    """
    from io import BytesIO
    from collections import OrderedDict
    from fastapi.responses import StreamingResponse
    from urllib.parse import quote
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    seller_type = (seller_type or "freelancer").lower()
    if seller_type not in ("freelancer", "business"):
        seller_type = "freelancer"

    # 파트너 조회 (삭제된 파트너도 export 허용)
    p_r = await db.execute(
        select(AffiliatePartner).where(AffiliatePartner.id == partner_id)
    )
    partner = p_r.scalar_one_or_none()
    if not partner:
        raise HTTPException(status_code=404, detail="Partner not found")

    # 기간 파싱
    range_start: Optional[datetime] = None
    range_end_excl: Optional[datetime] = None
    if start:
        try:
            range_start = datetime.strptime(start, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="start 형식: YYYY-MM-DD")
    if end:
        try:
            range_end_excl = datetime.strptime(end, "%Y-%m-%d") + timedelta(days=1)
        except ValueError:
            raise HTTPException(status_code=400, detail="end 형식: YYYY-MM-DD")

    # 전환 조회
    conv_query = select(ReferralConversion).where(
        ReferralConversion.partner_id == partner_id
    )
    if range_start is not None:
        conv_query = conv_query.where(ReferralConversion.converted_at >= range_start)
    if range_end_excl is not None:
        conv_query = conv_query.where(ReferralConversion.converted_at < range_end_excl)
    conv_query = conv_query.order_by(ReferralConversion.converted_at.desc())

    conv_r = await db.execute(conv_query)
    conversions = list(conv_r.scalars().all())

    # 캠페인 정보 캐시 (이름, 상품명, 수수료율)
    camp_ids = list({c.campaign_id for c in conversions if c.campaign_id})
    campaign_map: dict[int, dict] = {}
    if camp_ids:
        camp_r = await db.execute(
            select(AffiliateCampaign).where(AffiliateCampaign.id.in_(camp_ids))
        )
        for camp in camp_r.scalars().all():
            # 다중상품(카테고리 캠페인)이면 카테고리명, 단일이면 상품명
            product = (
                camp.cafe24_product_name
                or camp.cafe24_category_name
                or camp.product
                or camp.name
                or ""
            )
            campaign_map[camp.id] = {
                "name": camp.name or "",
                "product": product,
                "rate": float(camp.commission_rate or 0),
            }

    def _camp_info(cid: Optional[int]) -> dict:
        return campaign_map.get(cid or 0, {"name": "", "product": "", "rate": 0.0})

    STATUS_KOR = {"paid": "정상", "refunded": "환불", "cancelled": "취소"}

    def _fmt_dt(dt: Optional[datetime]) -> str:
        return dt.strftime("%Y-%m-%d %H:%M") if dt else ""

    # ─── 공통 스타일 ─────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2A2D35", end_color="2A2D35", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    bold = Font(bold=True)
    thin = Side(style="thin", color="888888")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt = "#,##0"
    pct_fmt = "0.0\"%\""

    wb = Workbook()

    # ===== Sheet 1: [요약] ===================================================
    ws_sum = wb.active
    ws_sum.title = "요약"

    # 제목
    title_text = (
        "인플루언서 정산서 (프리랜서)" if seller_type == "freelancer"
        else "인플루언서 정산서 (사업자)"
    )
    ws_sum["A1"] = title_text
    ws_sum["A1"].font = Font(bold=True, size=16)
    ws_sum.merge_cells("A1:G1")
    ws_sum["A1"].alignment = center

    # 메타 정보
    ws_sum["A3"] = "인플루언서명"
    ws_sum["B3"] = partner.name or ""
    ws_sum["A4"] = "정산 기간"
    ws_sum["B4"] = f"{start or '전체'} ~ {end or '전체'}"
    ws_sum["A5"] = "발행일"
    ws_sum["B5"] = datetime.utcnow().strftime("%Y-%m-%d")
    for r in (3, 4, 5):
        ws_sum[f"A{r}"].font = bold
        ws_sum[f"A{r}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # 매출요약 헤더
    # 프리랜서: 구분, 상품명, 판매수량, 총주문금액(원), 총공급가, 수수료(%), 크리에이터 정산금액
    # 사업자  : 구분, 상품명, 판매수량, 총주문금액(원), 수수료(%), 크리에이터 정산금액
    is_freelancer = seller_type == "freelancer"
    if is_freelancer:
        sum_headers = ["구분", "상품명", "판매수량", "총주문금액(원)", "총공급가(원)", "수수료(%)", "크리에이터 정산금액(원)"]
    else:
        sum_headers = ["구분", "상품명", "판매수량", "총주문금액(원)", "수수료(%)", "크리에이터 정산금액(원)"]

    summary_title_row = 7
    ws_sum.cell(row=summary_title_row, column=1, value="매출요약").font = Font(bold=True, size=12)

    header_row = summary_title_row + 1
    for ci, h in enumerate(sum_headers, start=1):
        cell = ws_sum.cell(row=header_row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = box

    # 정상/취소 → 상품별 그룹화 (OrderedDict로 입력 순서 보존)
    def _new_grp() -> dict:
        return {"qty": 0, "amount": 0.0, "rate": 0.0}

    normal_groups: "OrderedDict[str, dict]" = OrderedDict()
    cancel_groups: "OrderedDict[str, dict]" = OrderedDict()

    for c in conversions:
        info = _camp_info(c.campaign_id)
        product = info["product"] or "(상품명 미지정)"
        rate = info["rate"] or 0.0
        amount = float(c.order_amount or 0)
        qty = 1  # ReferralConversion에 수량 컬럼 없음 — 기본 1, 사용자가 엑셀에서 수정 가능
        is_cancel = (c.status in ("refunded", "cancelled")) or (c.refunded_amount or 0) > 0
        bucket = cancel_groups if is_cancel else normal_groups
        grp = bucket.setdefault(product, _new_grp())
        grp["qty"] += qty
        grp["amount"] += amount
        # rate는 상품별 첫 캠페인 rate (그룹 내 다른 캠페인 rate가 다르면 사용자가 수동 수정)
        if grp["rate"] == 0.0:
            grp["rate"] = rate

    row = header_row + 1
    section_start_rows: dict[str, tuple[int, int]] = {}  # label → (first, last)

    def _write_group(label: str, groups: "OrderedDict[str, dict]") -> Optional[tuple[int, int]]:
        """그룹 섹션 작성. 빈 그룹이면 placeholder 1줄. 첫/마지막 행 번호 리턴."""
        nonlocal row
        first = row
        if not groups:
            ws_sum.cell(row=row, column=1, value=label).alignment = center
            ws_sum.cell(row=row, column=2, value="(해당 주문 없음)")
            for ci in range(1, len(sum_headers) + 1):
                ws_sum.cell(row=row, column=ci).border = box
            row += 1
            return (first, row - 1)

        for product, g in groups.items():
            qty = g["qty"]
            amount = g["amount"]
            rate = g["rate"]
            ws_sum.cell(row=row, column=1, value=label).alignment = center
            ws_sum.cell(row=row, column=2, value=product)
            ws_sum.cell(row=row, column=3, value=qty)
            ws_sum.cell(row=row, column=4, value=amount).number_format = money_fmt
            if is_freelancer:
                # 총공급가 = 총주문금액 / 1.1, 정산금액 = 총공급가 * 수수료
                ws_sum.cell(row=row, column=5, value=f"=ROUND(D{row}/1.1, 0)").number_format = money_fmt
                ws_sum.cell(row=row, column=6, value=rate).number_format = pct_fmt
                ws_sum.cell(row=row, column=7, value=f"=ROUND(E{row}*F{row}/100, 0)").number_format = money_fmt
            else:
                # 사업자: 정산금액 = 총주문금액 * 수수료 (부가세 별도 처리)
                ws_sum.cell(row=row, column=5, value=rate).number_format = pct_fmt
                ws_sum.cell(row=row, column=6, value=f"=ROUND(D{row}*E{row}/100, 0)").number_format = money_fmt
            for ci in range(1, len(sum_headers) + 1):
                ws_sum.cell(row=row, column=ci).border = box
            row += 1
        return (first, row - 1)

    # 정상주문 섹션
    section_start_rows["normal"] = _write_group("정상주문", normal_groups)
    # 정상주문 소계
    normal_subtotal_row = row
    n_first, n_last = section_start_rows["normal"]
    ws_sum.cell(row=row, column=1, value="정상주문 소계").font = bold
    ws_sum.cell(row=row, column=1).alignment = center
    if normal_groups:
        ws_sum.cell(row=row, column=3, value=f"=SUM(C{n_first}:C{n_last})").number_format = money_fmt
        ws_sum.cell(row=row, column=4, value=f"=SUM(D{n_first}:D{n_last})").number_format = money_fmt
        if is_freelancer:
            ws_sum.cell(row=row, column=5, value=f"=SUM(E{n_first}:E{n_last})").number_format = money_fmt
            ws_sum.cell(row=row, column=7, value=f"=SUM(G{n_first}:G{n_last})").number_format = money_fmt
        else:
            ws_sum.cell(row=row, column=6, value=f"=SUM(F{n_first}:F{n_last})").number_format = money_fmt
    for ci in range(1, len(sum_headers) + 1):
        ws_sum.cell(row=row, column=ci).font = bold
        ws_sum.cell(row=row, column=ci).fill = PatternFill(start_color="FFF7E0", end_color="FFF7E0", fill_type="solid")
        ws_sum.cell(row=row, column=ci).border = box
    row += 1

    # 취소/환불 섹션
    section_start_rows["cancel"] = _write_group("취소/환불", cancel_groups)
    cancel_subtotal_row = row
    c_first, c_last = section_start_rows["cancel"]
    ws_sum.cell(row=row, column=1, value="취소/환불 소계").font = bold
    ws_sum.cell(row=row, column=1).alignment = center
    if cancel_groups:
        ws_sum.cell(row=row, column=3, value=f"=SUM(C{c_first}:C{c_last})").number_format = money_fmt
        ws_sum.cell(row=row, column=4, value=f"=SUM(D{c_first}:D{c_last})").number_format = money_fmt
        if is_freelancer:
            ws_sum.cell(row=row, column=5, value=f"=SUM(E{c_first}:E{c_last})").number_format = money_fmt
            ws_sum.cell(row=row, column=7, value=f"=SUM(G{c_first}:G{c_last})").number_format = money_fmt
        else:
            ws_sum.cell(row=row, column=6, value=f"=SUM(F{c_first}:F{c_last})").number_format = money_fmt
    for ci in range(1, len(sum_headers) + 1):
        ws_sum.cell(row=row, column=ci).font = bold
        ws_sum.cell(row=row, column=ci).fill = PatternFill(start_color="FDECEC", end_color="FDECEC", fill_type="solid")
        ws_sum.cell(row=row, column=ci).border = box
    row += 1

    # 합계 (정상+취소) — 판매수량·주문금액·공급가는 두 그룹 합산.
    # 크리에이터 정산금액은 전체주문건 시트의 [수수료(원)] 합계 행을 직접 참조해서
    # 사용자가 H열 수량을 수정해도 그 결과가 그대로 반영되도록 한다.
    detail_sum_row_idx = len(conversions) + 2  # 전체주문건 시트의 H열 합계 행 위치
    detail_commission_sum_ref = (
        f"'전체주문건'!H{detail_sum_row_idx}" if conversions else "0"
    )

    total_row = row
    ws_sum.cell(row=row, column=1, value="합계 (정상+취소)").font = bold
    ws_sum.cell(row=row, column=1).alignment = center
    ws_sum.cell(row=row, column=3, value=f"=C{normal_subtotal_row}+C{cancel_subtotal_row}").number_format = money_fmt
    ws_sum.cell(row=row, column=4, value=f"=D{normal_subtotal_row}+D{cancel_subtotal_row}").number_format = money_fmt
    if is_freelancer:
        ws_sum.cell(row=row, column=5, value=f"=E{normal_subtotal_row}+E{cancel_subtotal_row}").number_format = money_fmt
        ws_sum.cell(row=row, column=7, value=f"={detail_commission_sum_ref}").number_format = money_fmt
        final_settlement_cell = f"G{total_row}"
    else:
        ws_sum.cell(row=row, column=6, value=f"={detail_commission_sum_ref}").number_format = money_fmt
        final_settlement_cell = f"F{total_row}"
    for ci in range(1, len(sum_headers) + 1):
        ws_sum.cell(row=row, column=ci).font = Font(bold=True, size=12)
        ws_sum.cell(row=row, column=ci).fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        ws_sum.cell(row=row, column=ci).border = box
    row += 2  # 한 줄 띄움

    # ─── 정산 내역 ────────────────────────────────────────────────────────
    ws_sum.cell(row=row, column=1, value="정산 내역").font = Font(bold=True, size=12)
    row += 1
    detail_header_row = row
    for ci, h in enumerate(["항목", "금액(원)"], start=1):
        cell = ws_sum.cell(row=row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = box
    row += 1

    if is_freelancer:
        # 최종합계 = 크리에이터 정산금액 합계 (G{total_row})
        ws_sum.cell(row=row, column=1, value="최종합계 (크리에이터 정산금액)")
        ws_sum.cell(row=row, column=2, value=f"={final_settlement_cell}").number_format = money_fmt
        final_total_cell = f"B{row}"
        for ci in (1, 2):
            ws_sum.cell(row=row, column=ci).border = box
        row += 1
        # 소득세 3%
        ws_sum.cell(row=row, column=1, value="소득세 (3%)")
        ws_sum.cell(row=row, column=2, value=f"=ROUND({final_total_cell}*0.03, 0)").number_format = money_fmt
        income_tax_cell = f"B{row}"
        for ci in (1, 2):
            ws_sum.cell(row=row, column=ci).border = box
        row += 1
        # 주민세 0.3%
        ws_sum.cell(row=row, column=1, value="주민세 (0.3%)")
        ws_sum.cell(row=row, column=2, value=f"=ROUND({final_total_cell}*0.003, 0)").number_format = money_fmt
        local_tax_cell = f"B{row}"
        for ci in (1, 2):
            ws_sum.cell(row=row, column=ci).border = box
        row += 1
        # 실 지급 금액
        ws_sum.cell(row=row, column=1, value="실 지급 금액").font = Font(bold=True, size=12)
        ws_sum.cell(row=row, column=2, value=f"={final_total_cell}-{income_tax_cell}-{local_tax_cell}").number_format = money_fmt
        ws_sum.cell(row=row, column=2).font = Font(bold=True, size=12, color="C0392B")
        for ci in (1, 2):
            ws_sum.cell(row=row, column=ci).fill = PatternFill(start_color="FFF7E0", end_color="FFF7E0", fill_type="solid")
            ws_sum.cell(row=row, column=ci).border = box
        row += 1
    else:
        # 사업자: 최종합계 = 실지급 (세금계산서 본인 발행)
        ws_sum.cell(row=row, column=1, value="최종합계 (크리에이터 정산금액)")
        ws_sum.cell(row=row, column=2, value=f"={final_settlement_cell}").number_format = money_fmt
        final_total_cell = f"B{row}"
        for ci in (1, 2):
            ws_sum.cell(row=row, column=ci).border = box
        row += 1
        ws_sum.cell(row=row, column=1, value="실 지급 금액").font = Font(bold=True, size=12)
        ws_sum.cell(row=row, column=2, value=f"={final_total_cell}").number_format = money_fmt
        ws_sum.cell(row=row, column=2).font = Font(bold=True, size=12, color="C0392B")
        for ci in (1, 2):
            ws_sum.cell(row=row, column=ci).fill = PatternFill(start_color="FFF7E0", end_color="FFF7E0", fill_type="solid")
            ws_sum.cell(row=row, column=ci).border = box
        row += 1
        # 안내문
        row += 1
        note_cell = ws_sum.cell(row=row, column=1, value="※ 사업자 — 부가세 별도. 세금계산서는 인플루언서 본인이 발행합니다.")
        note_cell.font = Font(italic=True, color="666666", size=9)
        ws_sum.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(sum_headers))

    # 컬럼 폭
    summary_widths = [16, 28, 10, 16, 16, 12, 20]
    for i, w in enumerate(summary_widths[:len(sum_headers)], start=1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w
    ws_sum.freeze_panes = "A8"

    # ===== Sheet 2: [전체주문건] / Sheet 3: [취소건] =========================
    # 컬럼: A=주문번호 B=주문일시 C=상품명 D=캠페인 E=주문금액(원) F=주문수량
    #       G=수수료(%) H=수수료(원, 수식) I=환불금액(원) J=상태 K=환불일시
    DETAIL_HEADER = [
        "주문번호", "주문일시", "상품명", "캠페인",
        "주문금액(원)", "주문수량", "수수료(%)", "수수료(원)",
        "환불금액(원)", "상태", "환불일시",
    ]
    NUM_COLS = {"E": 5, "F": 6, "G": 7, "H": 8, "I": 9}

    def _detail_row(c: ReferralConversion) -> list:
        info = _camp_info(c.campaign_id)
        return [
            c.cafe24_order_id or c.order_id or f"#{c.id}",
            _fmt_dt(c.converted_at),
            info["product"] or "",
            info["name"] or "",
            int(c.order_amount or 0),
            1,                            # 주문수량 (기본 1, 사용자 수정 가능)
            info["rate"] or 0.0,          # 수수료(%) — 캠페인 commission_rate
            None,                         # 수수료(원) — 수식 (아래에서 채움)
            int(c.refunded_amount or 0),
            STATUS_KOR.get(c.status or "paid", c.status or "paid"),
            _fmt_dt(c.refunded_at),
        ]

    def _write_detail_sheet(ws, rows: list[list], title: str, summary_label: str):
        ws.title = title
        # 헤더
        for ci, h in enumerate(DETAIL_HEADER, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        # 데이터
        for ri, r in enumerate(rows, start=2):
            for ci, val in enumerate(r, start=1):
                ws.cell(row=ri, column=ci, value=val)
            # H{ri} 수수료(원) 수식 = ROUND(E*F*G/100, 0)
            ws.cell(row=ri, column=8, value=f"=ROUND(E{ri}*F{ri}*G{ri}/100, 0)")
            # 숫자/퍼센트 포맷
            for col in ("E", "F", "H", "I"):
                ws[f"{col}{ri}"].number_format = money_fmt
            ws[f"G{ri}"].number_format = pct_fmt
        # 합계 행 — SUM 수식
        if rows:
            sum_row = len(rows) + 2
            ws.cell(row=sum_row, column=1, value=summary_label).font = bold
            for col_letter in ("E", "F", "H", "I"):
                ws.cell(
                    row=sum_row,
                    column=NUM_COLS[col_letter],
                    value=f"=SUM({col_letter}2:{col_letter}{sum_row - 1})",
                ).number_format = money_fmt
                ws.cell(row=sum_row, column=NUM_COLS[col_letter]).font = bold
            # 합계 행 배경
            for ci in range(1, len(DETAIL_HEADER) + 1):
                ws.cell(row=sum_row, column=ci).fill = PatternFill(
                    start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                )
        # 컬럼 폭
        widths = [20, 17, 26, 22, 14, 10, 10, 14, 14, 10, 17]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    # 전체주문건 시트
    ws_all = wb.create_sheet("전체주문건")
    all_rows = [_detail_row(c) for c in conversions]
    _write_detail_sheet(ws_all, all_rows, "전체주문건", f"합계({len(all_rows)}건)")

    # 취소건 시트
    cancelled_convs = [
        c for c in conversions
        if (c.status in ("refunded", "cancelled")) or (c.refunded_amount or 0) > 0
    ]
    ws_cancel = wb.create_sheet("취소건")
    cancel_rows = [_detail_row(c) for c in cancelled_convs]
    _write_detail_sheet(ws_cancel, cancel_rows, "취소건", f"합계({len(cancel_rows)}건)")

    # 메모리 버퍼 출력
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    # 파일명
    today_str = datetime.utcnow().strftime("%Y%m%d")
    type_kor = "프리랜서" if is_freelancer else "사업자"
    range_part = ""
    if start and end:
        range_part = f"_{start.replace('-', '')}-{end.replace('-', '')}"
    elif start:
        range_part = f"_{start.replace('-', '')}_시작"
    elif end:
        range_part = f"_~{end.replace('-', '')}"
    safe_name = (partner.name or f"partner{partner_id}").replace("/", "_").replace("\\", "_")
    filename = f"정산서_{safe_name}_{type_kor}{range_part}_{today_str}.xlsx"
    quoted = quote(filename)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quoted}",
        },
    )


# ---------------------------------------------------------------------------
# Referral Programs
# ---------------------------------------------------------------------------

@router.post("/referral-programs", status_code=status.HTTP_201_CREATED)
async def create_referral_program(
    payload: ReferralProgramCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Create a new referral reward program."""
    program = ReferralProgram(
        user_id=current_user.id,
        **payload.model_dump(),
    )
    db.add(program)
    await db.commit()
    await db.refresh(program)
    return program


@router.get("/referral-programs")
async def list_referral_programs(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List referral programs (전체 관리자 공유)."""
    result = await db.execute(select(ReferralProgram))
    return result.scalars().all()


@router.put("/referral-programs/{program_id}")
async def update_referral_program(
    program_id: int,
    payload: ReferralProgramUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Update an existing referral program (전체 관리자 공유)."""
    result = await db.execute(
        select(ReferralProgram).where(ReferralProgram.id == program_id)
    )
    program = result.scalar_one_or_none()
    if not program:
        raise HTTPException(status_code=404, detail="Referral program not found")

    for field, value in payload.model_dump(exclude_none=True).items():
        setattr(program, field, value)

    await db.commit()
    await db.refresh(program)
    return program


# ---------------------------------------------------------------------------
# Click tracking
# ---------------------------------------------------------------------------

@router.get("/track/{referral_code}")
async def track_referral_click(
    referral_code: str,
    request: Request,
    db: AsyncSession = Depends(get_db),
):
    """
    클릭 추적 후 실제 Cafe24 스토어 상품 페이지로 리다이렉트.

    코드 우선순위:
    1. PartnerCampaign (파트너-캠페인별 코드)
    2. AffiliatePartner (파트너 기본 코드)
    3. AffiliateCampaign (캠페인 자체 코드, 파트너 없이 운영 가능)
    """
    partner_id: Optional[int] = None
    campaign_id: Optional[int] = None
    campaign: Optional[AffiliateCampaign] = None
    campaign_user: Optional[User] = None

    # 1) PartnerCampaign
    pc_result = await db.execute(
        select(PartnerCampaign).where(PartnerCampaign.referral_code == referral_code)
    )
    pc = pc_result.scalar_one_or_none()
    if pc:
        partner_id = pc.partner_id
        campaign_id = pc.campaign_id
        camp_result = await db.execute(
            select(AffiliateCampaign).where(AffiliateCampaign.id == pc.campaign_id)
        )
        campaign = camp_result.scalar_one_or_none()
    else:
        # 2) AffiliatePartner (휴지통 제외)
        result = await db.execute(
            select(AffiliatePartner).where(
                AffiliatePartner.referral_code == referral_code,
                AffiliatePartner.deleted_at.is_(None),
            )
        )
        partner = result.scalar_one_or_none()
        if partner:
            partner_id = partner.id
            campaign_id = partner.campaign_id
            if partner.campaign_id:
                camp_result = await db.execute(
                    select(AffiliateCampaign).where(AffiliateCampaign.id == partner.campaign_id)
                )
                campaign = camp_result.scalar_one_or_none()
        else:
            # 3) AffiliateCampaign 자체 코드
            camp_result = await db.execute(
                select(AffiliateCampaign).where(AffiliateCampaign.referral_code == referral_code)
            )
            campaign = camp_result.scalar_one_or_none()
            if campaign:
                campaign_id = campaign.id

    if not campaign and partner_id is None:
        raise HTTPException(status_code=404, detail="Invalid referral code")

    # 캠페인 소유자 조회 → public 도메인 결정
    if campaign:
        user_result = await db.execute(select(User).where(User.id == campaign.user_id))
        campaign_user = user_result.scalar_one_or_none()

    # 목적지 URL 동적 빌드 (저장된 값 대신 항상 최신 Cafe24 연결/쿠폰으로 생성)
    redirect_url = _build_destination_url(campaign, campaign_user) or "/"

    cookie_id = request.cookies.get("ref_id") or uuid.uuid4().hex
    ip_address = request.client.host if request.client else None
    user_agent = request.headers.get("user-agent")

    # partner_id가 None이면(캠페인 자체 클릭) 0 또는 skip
    if partner_id is not None:
        click = ReferralClick(
            partner_id=partner_id,
            campaign_id=campaign_id,
            ip_address=ip_address,
            user_agent=user_agent,
            cookie_id=cookie_id,
        )
        db.add(click)
        await db.commit()
        logger.info(f"[Track] click recorded: code={referral_code} partner={partner_id} campaign={campaign_id}")
    else:
        # 캠페인 자체 클릭 — partner_id NULL 허용하도록 로그만
        logger.info(f"[Track] campaign-only click: code={referral_code} campaign={campaign_id}")

    response = RedirectResponse(url=redirect_url, status_code=302)
    response.set_cookie(
        key="ref_id",
        value=cookie_id,
        max_age=60 * 60 * 24 * 30,
        httponly=True,
        samesite="lax",
    )
    return response


# ---------------------------------------------------------------------------
# Conversions (webhook)
# ---------------------------------------------------------------------------

@router.post("/conversions", status_code=status.HTTP_201_CREATED)
async def record_conversion(
    payload: ConversionCreate,
    db: AsyncSession = Depends(get_db),
):
    """
    Record a referral conversion.

    Intended as a webhook endpoint called by e-commerce platforms (e.g. Cafe24).
    No user auth required — validated by referral_code.
    """
    partner_result = await db.execute(
        select(AffiliatePartner).where(AffiliatePartner.referral_code == payload.referral_code)
    )
    partner = partner_result.scalar_one_or_none()
    if not partner:
        raise HTTPException(status_code=404, detail="Invalid referral code")

    # Find the most recent click matching the cookie_id if provided
    click_id: Optional[int] = None
    if payload.cookie_id:
        click_result = await db.execute(
            select(ReferralClick)
            .where(
                ReferralClick.partner_id == partner.id,
                ReferralClick.cookie_id == payload.cookie_id,
            )
            .order_by(ReferralClick.clicked_at.desc())
            .limit(1)
        )
        click = click_result.scalar_one_or_none()
        if click:
            click_id = click.id

    # Calculate commission
    commission_amount = 0.0
    if partner.campaign_id:
        camp_result = await db.execute(
            select(AffiliateCampaign).where(AffiliateCampaign.id == partner.campaign_id)
        )
        campaign = camp_result.scalar_one_or_none()
        if campaign:
            if campaign.commission_type == "percentage":
                commission_amount = payload.order_amount * (campaign.commission_rate / 100)
            else:
                commission_amount = campaign.commission_rate

    conversion = ReferralConversion(
        click_id=click_id,
        partner_id=partner.id,
        campaign_id=partner.campaign_id,
        order_id=payload.order_id,
        order_amount=payload.order_amount,
        commission_amount=round(commission_amount, 2),
    )
    db.add(conversion)
    await db.commit()
    await db.refresh(conversion)
    return {
        "success": True,
        "conversion_id": conversion.id,
        "commission_amount": conversion.commission_amount,
    }


# ---------------------------------------------------------------------------
# Phase 4 — 포인트 원장 & 내 추천 코드
# ---------------------------------------------------------------------------

@router.get("/my-points")
async def get_my_points(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """현재 유저의 포인트 잔액 + 최근 거래 50건."""
    balance_result = await db.execute(
        select(func.coalesce(func.sum(PointTransaction.amount), 0.0)).where(
            PointTransaction.user_id == current_user.id
        )
    )
    balance = float(balance_result.scalar())

    txn_result = await db.execute(
        select(PointTransaction)
        .where(PointTransaction.user_id == current_user.id)
        .order_by(PointTransaction.created_at.desc())
        .limit(50)
    )
    transactions = txn_result.scalars().all()
    return {
        "balance": balance,
        "transactions": [
            {
                "id": t.id,
                "amount": t.amount,
                "reason": t.reason,
                "memo": t.memo,
                "created_at": t.created_at,
            }
            for t in transactions
        ],
    }


@router.get("/my-referral-code")
async def get_my_referral_code(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """내 추천 코드 조회. 없으면 자동 생성."""
    if not current_user.referral_code:
        for _ in range(3):
            code = uuid.uuid4().hex[:8].upper()
            dup = await db.execute(select(User).where(User.referral_code == code))
            if not dup.scalar_one_or_none():
                current_user.referral_code = code
                await db.commit()
                break

    return {
        "referral_code": current_user.referral_code,
        "referral_link": f"{_settings.FRONTEND_URL}?ref={current_user.referral_code}",
    }


@router.post("/my-points/award", status_code=status.HTTP_201_CREATED)
async def award_points(
    payload: PointAwardRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """(관리자) 수동 포인트 지급."""
    if not current_user.is_superuser:
        raise HTTPException(status_code=403, detail="관리자 권한이 필요합니다.")

    txn = PointTransaction(
        user_id=payload.user_id,
        amount=payload.amount,
        reason=payload.reason,
        memo=payload.memo,
    )
    db.add(txn)
    await db.commit()
    await db.refresh(txn)
    return {"success": True, "transaction_id": txn.id}


# ---------------------------------------------------------------------------
# Phase 5 — 파트너 다캠페인 M:N 엔드포인트
# ---------------------------------------------------------------------------

@router.post("/partners/{partner_id}/campaigns", status_code=status.HTTP_201_CREATED)
async def add_partner_campaign(
    partner_id: int,
    payload: PartnerCampaignAdd,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """파트너에 캠페인 연결 추가 (전체 관리자 공유)."""
    p_result = await db.execute(
        select(AffiliatePartner).where(AffiliatePartner.id == partner_id)
    )
    partner = p_result.scalar_one_or_none()
    if not partner:
        raise HTTPException(status_code=404, detail="Partner not found")

    c_result = await db.execute(
        select(AffiliateCampaign).where(AffiliateCampaign.id == payload.campaign_id)
    )
    campaign = c_result.scalar_one_or_none()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    # 중복 체크
    dup = await db.execute(
        select(PartnerCampaign).where(
            PartnerCampaign.partner_id == partner_id,
            PartnerCampaign.campaign_id == payload.campaign_id,
        )
    )
    if dup.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="이미 연결된 캠페인입니다.")

    pc_code = await _unique_pc_code(db)
    pc_link = _build_referral_link(campaign, pc_code, current_user)
    pc = PartnerCampaign(
        partner_id=partner_id,
        campaign_id=payload.campaign_id,
        referral_code=pc_code,
        referral_link=pc_link,
    )
    db.add(pc)
    await db.commit()
    await db.refresh(pc)
    return {
        "id": pc.id,
        "partner_id": pc.partner_id,
        "campaign_id": pc.campaign_id,
        "referral_code": pc.referral_code,
        "referral_link": pc.referral_link,
        "created_at": pc.created_at,
    }


@router.delete("/partners/{partner_id}/campaigns/{pc_id}", status_code=status.HTTP_204_NO_CONTENT)
async def remove_partner_campaign(
    partner_id: int,
    pc_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """파트너-캠페인 연결 해제. pc_id=-1은 legacy campaign_id 처리."""
    p_result = await db.execute(
        select(AffiliatePartner).where(AffiliatePartner.id == partner_id)
    )
    partner = p_result.scalar_one_or_none()
    if not partner:
        raise HTTPException(status_code=404, detail="Partner not found")

    # pc_id = -1 → 가상 행(legacy campaign_id) 제거: partner.campaign_id NULL 처리
    if pc_id == -1:
        if partner.campaign_id is None:
            raise HTTPException(status_code=404, detail="legacy 캠페인 연결이 없습니다")
        partner.campaign_id = None
        await db.commit()
        return

    pc_result = await db.execute(
        select(PartnerCampaign).where(
            PartnerCampaign.id == pc_id,
            PartnerCampaign.partner_id == partner_id,
        )
    )
    pc = pc_result.scalar_one_or_none()
    if not pc:
        raise HTTPException(status_code=404, detail="PartnerCampaign not found")

    # legacy campaign_id가 동일하면 함께 NULL 처리 (중복 연결 방지)
    if partner.campaign_id == pc.campaign_id:
        partner.campaign_id = None

    await db.delete(pc)
    await db.commit()


@router.get("/partners/{partner_id}/performance")
async def get_partner_performance(
    partner_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """파트너별 캠페인 성과 집계 (전체 관리자 공유). PartnerCampaign + legacy campaign_id 둘 다 포함."""
    p_result = await db.execute(
        select(AffiliatePartner).where(AffiliatePartner.id == partner_id)
    )
    partner_obj = p_result.scalar_one_or_none()
    if not partner_obj:
        raise HTTPException(status_code=404, detail="Partner not found")

    # PartnerCampaign 목록
    pc_result = await db.execute(
        select(PartnerCampaign).where(PartnerCampaign.partner_id == partner_id)
    )
    pcs = list(pc_result.scalars().all())

    # legacy: PartnerCampaign에 없는 partner.campaign_id 도 추가
    pc_campaign_ids = {pc.campaign_id for pc in pcs}
    if partner_obj.campaign_id and partner_obj.campaign_id not in pc_campaign_ids:
        # 가상 PartnerCampaign row 생성 (DB에는 없지만 응답용)
        class _VirtualPC:
            id = -1
            partner_id = partner_obj.id
            campaign_id = partner_obj.campaign_id
            referral_code = partner_obj.referral_code
            referral_link = partner_obj.referral_link
        pcs.append(_VirtualPC())

    result_rows = []
    for pc in pcs:
        camp_result = await db.execute(
            select(AffiliateCampaign).where(AffiliateCampaign.id == pc.campaign_id)
        )
        campaign = camp_result.scalar_one_or_none()

        clicks_result = await db.execute(
            select(func.count(ReferralClick.id)).where(
                ReferralClick.partner_id == partner_id,
                ReferralClick.campaign_id == pc.campaign_id,
            )
        )
        clicks = clicks_result.scalar() or 0

        conv_result = await db.execute(
            select(
                func.count(ReferralConversion.id),
                func.coalesce(func.sum(ReferralConversion.order_amount), 0),
                func.coalesce(func.sum(ReferralConversion.commission_amount), 0),
            ).where(
                ReferralConversion.partner_id == partner_id,
                ReferralConversion.campaign_id == pc.campaign_id,
                ReferralConversion.status == "paid",
            )
        )
        conv_row = conv_result.one()

        # 링크는 항상 현재 캠페인 상태로 새로 계산 (DB에 저장된 stale 링크 무시)
        fresh_link = _build_referral_link(campaign, pc.referral_code, current_user) if campaign else pc.referral_link

        result_rows.append({
            "pc_id": pc.id,
            "campaign_id": pc.campaign_id,
            "campaign_name": campaign.name if campaign else "",
            "referral_code": pc.referral_code,
            "referral_link": fresh_link,
            "clicks": clicks,
            "conversions": conv_row[0] or 0,
            "sales": float(conv_row[1]),
            "commission": float(conv_row[2]),
        })

    return result_rows


@router.get("/campaigns/{campaign_id}/cafe24-debug")
async def cafe24_debug_campaign(
    campaign_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    캠페인 진단 — DB에 저장된 cafe24_* 필드 + 카페24 라이브 카테고리 정보 + 추천 조치.

    /r/{code} 가 홈으로 302되는 원인을 즉시 파악하기 위함.
    """
    import app.services.cafe24 as cafe24_svc

    r = await db.execute(
        select(AffiliateCampaign).where(AffiliateCampaign.id == campaign_id)
    )
    campaign = r.scalar_one_or_none()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    cafe24_user = await _resolve_cafe24_user(current_user, db)
    domain = _cafe24_store_domain(cafe24_user)

    db_state = {
        "id": campaign.id,
        "name": campaign.name,
        "referral_code": campaign.referral_code,
        "cafe24_product_no": campaign.cafe24_product_no,
        "cafe24_product_name": campaign.cafe24_product_name,
        "cafe24_category_no": campaign.cafe24_category_no,
        "cafe24_category_name": campaign.cafe24_category_name,
        "cafe24_category_url": campaign.cafe24_category_url,
        "cafe24_product_nos_raw": campaign.cafe24_product_nos,
        "cafe24_coupon_code": campaign.cafe24_coupon_code,
        "landing_url": campaign.landing_url,
        "base_product_url": campaign.base_product_url,
    }

    # 어떤 모드인지 판정
    if campaign.cafe24_category_no:
        mode = "category"
    elif campaign.cafe24_product_no:
        mode = "single_product"
    else:
        mode = "none_or_legacy"

    # 카페24 라이브 카테고리 상태 (cafe24_category_no 있을 때만)
    live_category = None
    live_category_products: list = []
    if campaign.cafe24_category_no:
        live_category = await cafe24_svc.get_category(
            cafe24_user, db, category_no=int(campaign.cafe24_category_no),
        )
        live_category_products = await cafe24_svc.list_category_products(
            cafe24_user, db, category_no=int(campaign.cafe24_category_no),
        )

    # /r/{code} 가 실제로 어디로 보낼지 시뮬레이션
    simulated_destination = _build_destination_url(campaign, cafe24_user)

    # 카테고리 URL 후보 여러 개를 직접 HEAD/GET으로 프로브 — 어떤 패턴이 동작하는지 확인
    storefront_probes = []
    if campaign.cafe24_category_no and domain:
        cat_no = int(campaign.cafe24_category_no)
        url_candidates = [
            f"https://{domain}/category/cat-no/{cat_no}/category.html",
            f"https://{domain}/product/list.html?cate_no={cat_no}",
            f"https://{domain}/category/{cat_no}/",
        ]
        for u in url_candidates:
            probe = await cafe24_svc.probe_category_url(u)
            storefront_probes.append({"url": u, **probe})

    # 추천 조치
    recommendation = None
    if mode == "category":
        if live_category and live_category.get("exists") is False:
            recommendation = (
                f"카페24에서 카테고리(no={campaign.cafe24_category_no})를 찾을 수 없습니다. "
                "캠페인을 삭제하고 새로 만드세요."
            )
        else:
            use_disp = live_category.get("use_display") if live_category else None
            access = live_category.get("access_authority") if live_category else None
            working_probe = next((p for p in storefront_probes if p.get("ok")), None)
            if working_probe:
                recommendation = (
                    f"동작하는 URL 패턴 발견: {working_probe['url']}. "
                    "이 URL로 cafe24_category_url을 갱신하세요."
                )
            elif use_disp != "T":
                recommendation = (
                    f"카테고리 use_display={use_disp}. POST /campaigns/{campaign.id}/republish-category로 T 갱신 필요."
                )
            elif access and access != "A":
                recommendation = (
                    f"카테고리 access_authority={access} (회원 등급 제한). "
                    "전체 공개로 변경하려면 카페24 관리자에서 '접근 권한'을 '전체회원'으로 설정하거나 "
                    "republish-category 재호출."
                )
            else:
                recommendation = (
                    "use_display=T, access_authority=A인데도 모든 URL 패턴이 홈으로 302됩니다. "
                    "카페24 관리자(상품→상품분류 관리)에서 카테고리 #" + str(campaign.cafe24_category_no) +
                    "의 '진열함', 'PC쇼핑몰 진열', '모바일쇼핑몰 진열'을 직접 확인하세요. "
                    "또는 카테고리 페이지에 진열할 상품이 0개일 수 있습니다."
                )
    elif mode == "none_or_legacy":
        recommendation = (
            "이 캠페인은 카페24 카테고리/상품이 연결돼 있지 않아 홈으로 폴백됩니다. "
            "캠페인을 삭제 후 '비공개 카테고리(다중 상품)' 모드로 다시 만드세요."
        )

    # DB의 cafe24_product_nos vs 카페24 라이브 카테고리 상품 비교
    expected_nos: list[int] = []
    if campaign.cafe24_product_nos:
        try:
            parsed = json.loads(campaign.cafe24_product_nos)
            if isinstance(parsed, list):
                expected_nos = [int(x) for x in parsed if x]
        except Exception:
            pass
    live_nos = [int(p.get("product_no")) for p in live_category_products if p.get("product_no")]
    missing_in_category = sorted(set(expected_nos) - set(live_nos))
    extra_in_category = sorted(set(live_nos) - set(expected_nos))

    # recommendation 보강: 상품이 0개거나 미스매치면 안내
    if mode == "category" and live_category and live_category.get("exists") is not False:
        if len(live_nos) == 0 and len(expected_nos) > 0:
            recommendation = (
                f"카테고리는 살아있지만 카페24에 묶인 상품이 0개입니다 "
                f"(DB에는 {len(expected_nos)}개 있어야 함). "
                f"POST /campaigns/{campaign.id}/reattach-products 로 재첨부 필요."
            )
        elif missing_in_category:
            recommendation = (
                f"누락된 상품 {len(missing_in_category)}개: {missing_in_category[:5]}{'...' if len(missing_in_category) > 5 else ''}. "
                f"POST /campaigns/{campaign.id}/reattach-products 로 재첨부."
            )

    return {
        "mode": mode,
        "domain": domain,
        "db_state": db_state,
        "live_category": live_category,
        "live_category_products": live_category_products,
        "expected_product_nos": expected_nos,
        "live_product_nos": live_nos,
        "missing_in_category": missing_in_category,
        "extra_in_category": extra_in_category,
        "storefront_probes": storefront_probes,
        "simulated_destination": simulated_destination,
        "recommendation": recommendation,
    }


@router.post("/campaigns/{campaign_id}/reattach-products")
async def reattach_campaign_products(
    campaign_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    캠페인의 cafe24_product_nos를 카페24 카테고리에 재첨부.

    초기 attach_products_to_category가 잘못된 필드명(sort_no)으로 0개 첨부된
    카테고리를 복구할 때 사용. 이미 카테고리에 있는 상품은 카페24가 알아서 무시함.
    """
    import app.services.cafe24 as cafe24_svc

    r = await db.execute(
        select(AffiliateCampaign).where(AffiliateCampaign.id == campaign_id)
    )
    campaign = r.scalar_one_or_none()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if not campaign.cafe24_category_no:
        raise HTTPException(status_code=400, detail="이 캠페인은 카테고리에 연결돼 있지 않습니다.")
    if not campaign.cafe24_product_nos:
        raise HTTPException(status_code=400, detail="DB에 첨부할 상품 목록(cafe24_product_nos)이 없습니다.")

    try:
        product_nos = json.loads(campaign.cafe24_product_nos)
        if not isinstance(product_nos, list) or not product_nos:
            raise ValueError("empty list")
    except Exception:
        raise HTTPException(status_code=400, detail="cafe24_product_nos JSON 파싱 실패")

    cafe24_user = await _resolve_cafe24_user(current_user, db)
    try:
        result = await cafe24_svc.attach_products_to_category(
            cafe24_user, db,
            category_no=int(campaign.cafe24_category_no),
            product_nos=[int(p) for p in product_nos],
        )
    except Exception as e:
        logger.error(f"[Reattach] failed: {e}")
        raise HTTPException(status_code=502, detail=f"카페24 상품 첨부 실패: {str(e)[:200]}")

    # 첨부 후 실제 카테고리에 어떤 상품이 들어갔는지 검증
    live_after = await cafe24_svc.list_category_products(
        cafe24_user, db, category_no=int(campaign.cafe24_category_no),
    )
    errors = (result or {}).get("errors") or []
    return {
        "success": True,
        "category_no": campaign.cafe24_category_no,
        "expected_count": len(product_nos),
        "attached_count": (result or {}).get("attached", 0),
        "live_count_after": len(live_after),
        "errors": errors,
        "attach_result": result,
    }


@router.post("/campaigns/{campaign_id}/republish-category")
async def republish_campaign_category(
    campaign_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    기존 캠페인의 카페24 비공개 카테고리를 URL 접근 가능 상태로 전환.

    Phase 6 초기 버전에서 display_pc_yn=F로 만들어진 카테고리는 카페24가 홈으로 302시킴.
    이 엔드포인트가 PUT /categories/{N}으로 display_pc_yn=T, use_main_category=F로 갱신해
    "메뉴엔 안 보이지만 URL은 동작" 상태로 만든다.
    """
    import app.services.cafe24 as cafe24_svc

    r = await db.execute(
        select(AffiliateCampaign).where(AffiliateCampaign.id == campaign_id)
    )
    campaign = r.scalar_one_or_none()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if not campaign.cafe24_category_no:
        raise HTTPException(
            status_code=400,
            detail="이 캠페인에는 카페24 카테고리가 연결돼 있지 않습니다.",
        )

    cafe24_user = await _resolve_cafe24_user(current_user, db)
    domain = _cafe24_store_domain(cafe24_user)

    # 카테고리 진열/접근 권한 갱신 (best-effort)
    update_result = None
    update_error = None
    try:
        update_result = await cafe24_svc.update_category_visibility(
            cafe24_user, db,
            category_no=int(campaign.cafe24_category_no),
            display=True,
            use_main=False,
        )
    except Exception as e:
        logger.warning(f"[Republish] category {campaign.cafe24_category_no} update failed: {e}")
        update_error = str(e)[:200]

    # DB의 cafe24_category_url을 동작하는 새 URL 패턴으로 갱신
    new_url = cafe24_svc.category_storefront_url(domain, int(campaign.cafe24_category_no))
    campaign.cafe24_category_url = new_url
    campaign.base_product_url = new_url
    if not campaign.landing_url or "/category/cat-no/" in (campaign.landing_url or ""):
        campaign.landing_url = new_url
    await db.commit()
    await db.refresh(campaign)

    logger.info(
        f"[Republish] campaign={campaign_id} category={campaign.cafe24_category_no} "
        f"-> updated url={new_url} use_display={(update_result or {}).get('use_display')}"
    )
    return {
        "success": True,
        "campaign_id": campaign_id,
        "category_no": campaign.cafe24_category_no,
        "category_url": new_url,
        "update_result": update_result,
        "update_error": update_error,
    }


@router.get("/partners/{partner_id}/timeseries")
async def get_partner_timeseries(
    partner_id: int,
    days: int = Query(30, ge=1, le=365),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    파트너 일별 성과 시계열 — 일별 매출/전환/환불/취소/클릭/커미션 집계.

    카페24 실주문 일자(converted_at) 기준. 데이터가 없는 날도 0으로 채워서 전기간 반환.
    파트너 상세 모달의 일별 매출 트래킹 차트/로그용.
    """
    # 파트너 존재 검증
    p_r = await db.execute(
        select(AffiliatePartner).where(AffiliatePartner.id == partner_id)
    )
    if not p_r.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Partner not found")

    today = datetime.utcnow().date()
    start_date = today - timedelta(days=days - 1)
    start_dt = datetime.combine(start_date, datetime.min.time())

    # 일별 클릭
    click_rows = await db.execute(
        select(
            func.date(ReferralClick.clicked_at).label("d"),
            func.count(ReferralClick.id).label("clicks"),
        )
        .where(
            ReferralClick.partner_id == partner_id,
            ReferralClick.clicked_at >= start_dt,
        )
        .group_by(func.date(ReferralClick.clicked_at))
    )
    clicks_map: dict[str, int] = {}
    for row in click_rows.all():
        d = row[0]
        key = d.isoformat() if hasattr(d, "isoformat") else str(d)
        clicks_map[key] = int(row[1])

    # 일별 status별 집계 (paid/refunded/cancelled 모두 별개로)
    conv_rows = await db.execute(
        select(
            func.date(ReferralConversion.converted_at).label("d"),
            ReferralConversion.status,
            func.count(ReferralConversion.id).label("cnt"),
            func.coalesce(func.sum(ReferralConversion.order_amount), 0).label("amt"),
            func.coalesce(func.sum(ReferralConversion.commission_amount), 0).label("comm"),
        )
        .where(
            ReferralConversion.partner_id == partner_id,
            ReferralConversion.converted_at >= start_dt,
        )
        .group_by(func.date(ReferralConversion.converted_at), ReferralConversion.status)
    )
    conv_map: dict[str, dict] = {}
    for row in conv_rows.all():
        d = row[0]
        key = d.isoformat() if hasattr(d, "isoformat") else str(d)
        st = (row[1] or "").strip().lower()
        cnt = int(row[2])
        amt = float(row[3])
        comm = float(row[4])
        bucket = conv_map.setdefault(key, {
            "conversions": 0, "sales": 0.0, "commission": 0.0,
            "refunded_count": 0, "refunded_amount": 0.0,
            "cancelled_count": 0, "cancelled_amount": 0.0,
        })
        if st == "paid":
            bucket["conversions"] = cnt
            bucket["sales"] = amt
            bucket["commission"] = comm
        elif st == "refunded":
            bucket["refunded_count"] = cnt
            bucket["refunded_amount"] = amt
        elif st == "cancelled":
            bucket["cancelled_count"] = cnt
            bucket["cancelled_amount"] = amt

    # 모든 날짜를 채워서 반환
    series = []
    for i in range(days):
        d = (start_date + timedelta(days=i)).isoformat()
        c = conv_map.get(d, {})
        series.append({
            "date": d,
            "clicks": clicks_map.get(d, 0),
            "conversions": c.get("conversions", 0),
            "sales": c.get("sales", 0.0),
            "commission": c.get("commission", 0.0),
            "refunded_count": c.get("refunded_count", 0),
            "refunded_amount": c.get("refunded_amount", 0.0),
            "cancelled_count": c.get("cancelled_count", 0),
            "cancelled_amount": c.get("cancelled_amount", 0.0),
        })

    return series


@router.get("/partners/{partner_id}/audit")
async def audit_partner_conversions(
    partner_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    파트너 매출 정합성 진단 — status별 집계 + 원시 conversion 목록 반환.
    관리자 화면 vs 파트너 포털 매출 불일치 시 어느 쪽 데이터에 문제가 있는지 추적.
    """
    p_r = await db.execute(
        select(AffiliatePartner).where(AffiliatePartner.id == partner_id)
    )
    partner = p_r.scalar_one_or_none()
    if not partner:
        raise HTTPException(status_code=404, detail="Partner not found")

    # status별 그룹 집계 (관리자 기준과 동일)
    status_r = await db.execute(
        select(
            ReferralConversion.status,
            func.count(ReferralConversion.id),
            func.coalesce(func.sum(ReferralConversion.order_amount), 0),
            func.coalesce(func.sum(ReferralConversion.commission_amount), 0),
        ).where(ReferralConversion.partner_id == partner_id)
        .group_by(ReferralConversion.status)
    )
    status_breakdown = []
    paid_amount = 0.0
    gross_amount = 0.0
    for row in status_r.all():
        st_raw = row[0]
        amt = float(row[2])
        gross_amount += amt
        if (st_raw or "").strip().lower() == "paid":
            paid_amount = amt
        status_breakdown.append({
            "status_raw": st_raw,
            "status_normalized": (st_raw or "").strip().lower() or "(empty)",
            "count": int(row[1]),
            "order_amount_sum": amt,
            "commission_sum": float(row[3]),
        })

    # 원시 conversion 목록 (최근 200건)
    rows_r = await db.execute(
        select(ReferralConversion)
        .where(ReferralConversion.partner_id == partner_id)
        .order_by(ReferralConversion.id.desc())
        .limit(200)
    )
    conversions = [
        {
            "id": c.id,
            "campaign_id": c.campaign_id,
            "order_id": c.order_id,
            "cafe24_order_id": c.cafe24_order_id,
            "order_amount": c.order_amount,
            "commission_amount": c.commission_amount,
            "status": c.status,
            "refunded_amount": c.refunded_amount,
            "refunded_at": c.refunded_at.isoformat() if c.refunded_at else None,
            "converted_at": c.converted_at.isoformat() if c.converted_at else None,
        }
        for c in rows_r.scalars().all()
    ]

    return {
        "partner": {
            "id": partner.id,
            "name": partner.name,
            "phone": partner.phone,
            "email": partner.email,
        },
        "summary": {
            "net_sales_paid_only": paid_amount,
            "gross_sales_all_status": gross_amount,
            "diff": gross_amount - paid_amount,
        },
        "status_breakdown": status_breakdown,
        "conversions_recent_200": conversions,
    }
